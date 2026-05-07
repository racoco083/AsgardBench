# pylint: disable=missing-module-docstring,missing-class-docstring,missing-function-docstring
"""
Standalone script for generating test result reports.

This script scans the test directory structure and generates:
1. results.csv/xlsx - Aggregated results across all tests and models
2. plan_performance.csv/xlsx - Performance statistics for each unique plan

Usage:
    python -m AsgardBench.Model.generate_reports
    python AsgardBench/Model/generate_reports.py
"""

import argparse
import os
import sys
from typing import Any, List, Set, Tuple

from AsgardBench.Utils.json_utils import FileReadError, read_json_file


def print_error(message: str) -> None:
    """Print an error message in red to stderr."""
    RED = "\033[91m"
    RESET = "\033[0m"
    print(f"{RED}ERROR: {message}{RESET}", file=sys.stderr)


from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from AsgardBench import constants as c
from AsgardBench.Model.test_results import ResultStat, TestResults
from AsgardBench.Utils.config_utils import EvaluationConfig

# Baseline configuration string for comparing test results
BASELINE_CONFIG = "T0_Fs_H60_C0_P2_I1_R1_S1_E0_M4096"


class ResultsPrinter:
    """Handles discovery and printing of test results."""

    def __init__(self, test_dir: str | None = None):
        """Initialize the results printer.

        Args:
            test_dir: Directory containing test results. Defaults to c.TEST_DIR.
        """
        self.test_dir = test_dir or c.TEST_DIR
        # Cache for JSON files to avoid re-reading from blob storage
        self._json_cache: dict[str, Any] = {}

    def _read_json_cached(self, file_path: str) -> Any:
        """Read a JSON file with caching to avoid repeated blob storage reads.

        Args:
            file_path: Path to the JSON file.

        Returns:
            Parsed JSON data.
        """
        if file_path not in self._json_cache:
            self._json_cache[file_path] = read_json_file(file_path)
        return self._json_cache[file_path]

    def clear_cache(self):
        """Clear the JSON file cache."""
        self._json_cache.clear()

    def discover_tests_and_models(self) -> Tuple[List[str], List[str]]:
        """Discover available test names and model directories by scanning directory structure.

        Scans: {test_dir}/{test_name}/{model_dir}/config.json

        Returns:
            (model_dirs, tests) as sorted lists of unique names
        """
        tests: Set[str] = set()
        model_dirs: Set[str] = set()

        if not os.path.exists(self.test_dir):
            print(f"Test directory not found: {self.test_dir}")
            return [], []

        for test_name in os.listdir(self.test_dir):
            test_path = os.path.join(self.test_dir, test_name)
            if not os.path.isdir(test_path):
                continue
            tests.add(test_name)
            try:
                for entry in os.listdir(test_path):
                    entry_path = os.path.join(test_path, entry)
                    if not os.path.isdir(entry_path):
                        continue
                    # Direct match: {test}/{model_dir}/config.json
                    if os.path.exists(os.path.join(entry_path, "config.json")):
                        model_dirs.add(entry)
                    else:
                        # One level deeper: {test}/{prefix}/{model_dir}/config.json
                        # Handles model IDs with slashes like "qwen/qwen3.5-9b"
                        try:
                            for sub_entry in os.listdir(entry_path):
                                sub_path = os.path.join(entry_path, sub_entry)
                                if os.path.isdir(sub_path) and os.path.exists(
                                    os.path.join(sub_path, "config.json")
                                ):
                                    model_dirs.add(f"{entry}/{sub_entry}")
                        except PermissionError:
                            continue
            except PermissionError:
                continue

        sorted_model_dirs = sorted(model_dirs)
        sorted_tests = sorted(tests)

        print(
            f"Discovered tests ({len(sorted_tests)}): "
            + (", ".join(sorted_tests) if sorted_tests else "<none>")
        )
        print(
            f"Discovered model directories ({len(sorted_model_dirs)}): "
            + (", ".join(sorted_model_dirs) if sorted_model_dirs else "<none>")
        )

        return sorted_model_dirs, sorted_tests

    def collect_config_keys(self, test_list: List[str]) -> List[str]:
        """Collect all unique config keys from all config.json files.

        Args:
            test_list: List of test names to scan.

        Returns:
            Sorted list of unique config keys.
        """
        all_config_keys: Set[str] = set()

        for test_name in test_list:
            test_path = os.path.join(self.test_dir, test_name)
            if not os.path.isdir(test_path):
                continue

            for model_dir in os.listdir(test_path):
                config_file = os.path.join(test_path, model_dir, "config.json")
                if os.path.exists(config_file):
                    config_data = self._read_json_cached(config_file)
                    # Normalize config to handle backwards compatibility
                    config_data = EvaluationConfig.normalize_config_dict(config_data)
                    all_config_keys.update(config_data.keys())

        config_keys = sorted(all_config_keys)
        print(f"Found {len(config_keys)} config keys: {', '.join(config_keys)}")
        return config_keys

    def print_results_for_model(
        self,
        test_set_dirs: List[str],
        model_dir: str,
        file=None,
        config_keys: List[str] | None = None,
    ) -> Tuple[TestResults, dict | None]:
        """Print combined results from multiple test sets for a single model directory.

        Args:
            test_set_dirs: List of test set directory names.
            model_dir: Model directory name (as it appears in the filesystem).
            file: File handle to write CSV to.
            config_keys: List of config keys to include in CSV output.

        Returns:
            Tuple of (TestResults, config_data dict or None)
        """
        # Use "ALL" when combining multiple test sets, otherwise use the single test name
        test_set_names = "ALL" if len(test_set_dirs) > 1 else test_set_dirs[0]
        combined_results = TestResults(
            test_name=test_set_names, model_name=model_dir, expected_num_plans=0
        )
        combined_config: dict | None = None

        for test_set_dir in test_set_dirs:
            result_dir = f"{self.test_dir}/{test_set_dir}/{model_dir}"

            test_results_file = f"{result_dir}/test_results.json"
            config_file = f"{result_dir}/config.json"

            # Skip directories without config.json
            if not os.path.exists(config_file):
                print(f"Skipping {test_set_dir}/{model_dir} - no config.json found")
                continue

            if not os.path.exists(test_results_file):
                print(
                    f"Warning: No results found for {test_set_dir} with model {model_dir}"
                )
                continue

            # Load config data
            config_data = self._read_json_cached(config_file)
            # Normalize config to handle backwards compatibility
            config_data = EvaluationConfig.normalize_config_dict(config_data)

            # Use first config as the combined config
            if combined_config is None:
                combined_config = config_data

            results_json = self._read_json_cached(test_results_file)
            test_results = TestResults.from_dict(results_json)

            # Add all test results to combined results
            combined_results.test_results.extend(test_results.test_results)
            combined_results.expected_num_plans += test_results.expected_num_plans
            print(
                f"Loaded {len(test_results.test_results)} results from {test_set_dir}"
            )

        if combined_results.test_results:
            combined_results.print()
            combined_results.add_to_csv(file, combined_config, config_keys)
        else:
            print(f"No results found to display for {model_dir}.")

        return combined_results, combined_config

    def collect_results_row(
        self,
        combined_results: TestResults,
        combined_config: dict | None,
        config_keys: List[str],
    ) -> List[Any] | None:
        """Collect a single row of results data.

        Returns:
            List of values for the row, or None if no results.
        """
        if not combined_results.test_results:
            return None

        # Calculate statistics
        total_results = ResultStat()
        for test_result in combined_results.test_results:
            total_results.add_result(test_result)

        total = total_results.success_count + total_results.fail_count
        per_success = total_results.success_count / total if total > 0 else 0
        per_fail = total_results.fail_count / total if total > 0 else 0

        from AsgardBench.Model.test_results import FailType, StepExtension

        max_repeats = total_results.fail_reason_counts.get(FailType.MAX_REPEATS, 0)
        max_steps = total_results.fail_reason_counts.get(FailType.MAX_STEPS, 0)
        max_failures = total_results.fail_reason_counts.get(FailType.MAX_FAILURES, 0)
        fail_api = total_results.fail_reason_counts.get(FailType.API_FAILURE, 0)

        # Get step extension counts
        ext_none = total_results.step_extension_counts.get(StepExtension.NONE, 0)
        ext_success = total_results.step_extension_counts.get(StepExtension.EXTENDED, 0)
        ext_failure = total_results.step_extension_counts.get(
            StepExtension.HIT_HARD_LIMIT, 0
        )

        # Calculate percentages for max_repeats, max_steps, max_failures, fail_api
        # Store as decimals (0.0 to 1.0) for consistent Excel formatting with other percentages
        max_repeats_pct = (max_repeats / total) if total > 0 else 0
        max_steps_pct = (max_steps / total) if total > 0 else 0
        max_failures_pct = (max_failures / total) if total > 0 else 0
        fail_api_pct = (fail_api / total) if total > 0 else 0

        # Calculate percentages for step extension counts
        ext_none_pct = (ext_none / total) if total > 0 else 0
        ext_success_pct = (ext_success / total) if total > 0 else 0
        ext_failure_pct = (ext_failure / total) if total > 0 else 0

        # Calculate completion: True if complete, otherwise percentage (0.0 to 1.0)
        if combined_results.expected_num_plans > 0:
            if combined_results.expected_num_plans == total:
                test_completed = True
            else:
                test_completed = total / combined_results.expected_num_plans
        else:
            test_completed = True  # No plans expected = complete

        # Build row: test_completed first, then model_name, test_set_name, then other config values, then metrics
        row = []

        # Add test_completed as first column (True or percentage)
        row.append(test_completed)

        # Split model_name into model and config parts (separated by "--")
        model_name = ""
        if combined_config:
            model_name = combined_config.get("model_name", "")
        if "--" in model_name:
            model_part, config_part = model_name.split("--", 1)
        else:
            model_part = model_name
            config_part = ""
        row.append(model_part)
        row.append(config_part)

        # Add rep column (empty by default, filled in by print_all_results for rep runs)
        row.append("")

        # Add rep_count column (empty by default, filled in by print_all_results for AVG rows)
        row.append("")

        # Add test_set_name as sixth column
        row.append(combined_results.test_set_name)

        # Keys to move to the end (before git_commit)
        end_keys = {
            "max_completion_tokens",
            "prompt_version",
            "temperature",
            "implementation",
        }

        # Add other config values (excluding special ones)
        if combined_config and config_keys:
            for key in config_keys:
                if key in (
                    "model_name",
                    "test_set_name",
                    "test_name",
                    "git_commit",
                    "include_common_sense",
                ):
                    continue
                if key in end_keys:
                    continue
                row.append(combined_config.get(key, ""))

        # Add metrics
        row.extend(
            [
                per_success,
                per_fail,
                "",  # % change - calculated in write_xlsx
                "",  # success range - only for AVG+ALL rows
                "",  # success stdev - only for AVG+ALL rows
                max_repeats_pct,
                max_steps_pct,
                max_failures_pct,
                fail_api_pct,
                ext_none_pct,
                ext_success_pct,
                ext_failure_pct,
                total_results.success_count,
                total_results.fail_count,
                max_repeats,
                max_steps,
                max_failures,
                fail_api,
                ext_none,
                ext_success,
                ext_failure,
                total_results.step_ratio,
                total_results.valid_ratio,
                total_results.invalid_action_ratio,
                total_results.invalid_object_ratio,
                total_results.undoable_ratio,
                total_results.invalid_response_ratio,
                total_results.percent_goals_reached,
                combined_results.expected_num_plans,
            ]
        )

        # Add max_completion_tokens, prompt_version, temperature, implementation at the end
        if combined_config and config_keys:
            for key in [
                "max_completion_tokens",
                "prompt_version",
                "temperature",
                "implementation",
            ]:
                if key in config_keys:
                    row.append(combined_config.get(key, ""))

        # Add git_commit as last column
        if combined_config and "git_commit" in config_keys:
            row.append(combined_config.get("git_commit", ""))

        return row

    def get_headers(self, config_keys: List[str]) -> List[str]:
        """Get the header row."""
        # test_completed first, then model, config, rep, rep_count, test_set_name
        headers = [
            "test_completed",
            "model",
            "config",
            "rep",
            "rep_count",
            "test_set_name",
        ]

        # Keys to move to the end (before git_commit)
        end_keys = {
            "max_completion_tokens",
            "prompt_version",
            "temperature",
            "implementation",
        }

        # Add other config keys (excluding special ones)
        for key in config_keys:
            if key in (
                "model_name",
                "test_set_name",
                "test_name",
                "git_commit",
                "include_common_sense",
            ):
                continue
            if key in end_keys:
                continue
            headers.append(key)

        # Add metric headers
        headers.extend(
            [
                "success %",
                "fail %",
                "% change",
                "success range",
                "success stdev",
                "max repeat %",
                "max steps %",
                "max failures %",
                "fail api %",
                "ext none %",
                "ext success %",
                "ext failure %",
                "success #",
                "fail #",
                "max repeat #",
                "max steps #",
                "max failures #",
                "fail api #",
                "ext none #",
                "ext success #",
                "ext failure #",
                "step %",
                "valid %",
                "invalid action %",
                "invalid object %",
                "undoable %",
                "unparsable %",
                "goals reached %",
                "expected_num_plans",
            ]
        )

        # Add max_completion_tokens, prompt_version, temperature, implementation at the end
        for key in [
            "max_completion_tokens",
            "prompt_version",
            "temperature",
            "implementation",
        ]:
            if key in config_keys:
                headers.append(key)

        # Add git_commit as last column
        if "git_commit" in config_keys:
            headers.append("git_commit")

        return headers

    def write_xlsx(
        self,
        output_file: str,
        headers: List[str],
        rows: List[List[Any]],
        config_keys: List[str],
        workbook: Workbook | None = None,
        sheet_name: str = "Results",
    ) -> Workbook:
        """Write results to an Excel file with formatting.

        Args:
            output_file: Path to output xlsx file.
            headers: List of header names.
            rows: List of data rows.
            config_keys: List of config keys (for determining column count).
            workbook: Optional existing workbook to add sheet to.
            sheet_name: Name for the worksheet.

        Returns:
            The workbook (either newly created or the one passed in).
        """
        if workbook is None:
            wb = Workbook()
            ws = wb.active
            ws.title = sheet_name
        else:
            wb = workbook
            # If workbook already has this sheet, use it; otherwise create new one
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
            else:
                ws = wb.create_sheet(sheet_name)

        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(
            start_color="4472C4", end_color="4472C4", fill_type="solid"
        )
        header_alignment = Alignment(
            horizontal="center", vertical="bottom", wrap_text=False, textRotation=90
        )

        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # Border with bold right side for column group separators
        bold_right_border = Border(
            left=Side(style="thin"),
            right=Side(style="medium"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # Find column indices by header name for special formatting
        percentage_col_names = {
            "success %",
            "fail %",
            "success stdev",
            "max repeat %",
            "max steps %",
            "max failures %",
            "fail api %",
            "ext none %",
            "ext success %",
            "ext failure %",
            "step %",
            "valid %",
            "invalid action %",
            "invalid object %",
            "undoable %",
            "unparsable %",
            "goals reached %",
        }

        # Icon columns that display single characters (should have fixed narrow width)
        icon_col_names = {
            "hand_transparency",
            "feedback_type",
            "previous_image",
            "use_memory",
        }

        # Build sets of column indices (1-indexed)
        percentage_cols = set()
        integer_count_cols = set()  # Columns that should display as integers (rounded)
        icon_cols = set()  # Columns that display icons (fixed narrow width)
        model_path_col = None
        rep_col = None
        test_completed_col = None
        success_percentage_col = None
        fail_percentage_col = None
        fail_api_percentage_col = None
        hand_transparency_col = None
        feedback_type_col = None
        previous_image_col = None

        # Columns after which to add a bold right border
        bold_border_after_cols = {
            "use_memory",
            "fail %",
            "fail api %",
            "ext failure %",
            "fail #",
            "fail api #",
            "ext failure #",
            "step %",
            "unparsable %",
        }
        bold_border_col_indices = set()

        # Column names that should be displayed as integers (for AVG rows)
        integer_count_col_names = {
            "success #",
            "fail #",
            "max repeat #",
            "max steps #",
            "max failures #",
            "fail api #",
            "ext none #",
            "ext success #",
            "ext failure #",
            "expected_num_plans",
            "max_completion_tokens",
        }

        for col_idx, header in enumerate(headers, start=1):
            if header in percentage_col_names:
                percentage_cols.add(col_idx)
            if header in integer_count_col_names:
                integer_count_cols.add(col_idx)
            if header in icon_col_names:
                icon_cols.add(col_idx)
            if header in bold_border_after_cols:
                bold_border_col_indices.add(col_idx)
            if header == "model_path":
                model_path_col = col_idx
            elif header == "rep":
                rep_col = col_idx
            elif header == "test_completed":
                test_completed_col = col_idx
            elif header == "success %":
                success_percentage_col = col_idx
            elif header == "fail %":
                fail_percentage_col = col_idx
            elif header == "fail api %":
                fail_api_percentage_col = col_idx
            elif header == "hand_transparency":
                hand_transparency_col = col_idx
            elif header == "feedback_type":
                feedback_type_col = col_idx
            elif header == "previous_image":
                previous_image_col = col_idx

        # Write headers
        for col_idx, header in enumerate(headers, start=1):
            # Add leading space and replace underscores with spaces for readability
            display_header = header.replace("_", " ")
            # Rename hand_transparency to "show hand (60)"
            if header == "hand_transparency":
                display_header = "show hand (60)"
            display_header = "        " + display_header
            cell = ws.cell(row=1, column=col_idx, value=display_header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            # Apply bold right border for column group separators
            if col_idx in bold_border_col_indices:
                cell.border = bold_right_border
            else:
                cell.border = thin_border

        # Define fonts for boolean-like values and special columns
        green_font = Font(color="008000")  # Green for true/yes
        red_font = Font(color="FF0000")  # Red for false/no/none
        success_font = Font(
            bold=True, color="006400"
        )  # Dark green bold for success_percentage
        fail_font = Font(bold=True, color="8B0000")  # Dark red bold for fail_percentage

        # Define fill for incomplete tests
        light_red_fill = PatternFill(
            start_color="FFCCCC", end_color="FFCCCC", fill_type="solid"
        )
        # Define fill for rep rows (light blue)
        light_blue_fill = PatternFill(
            start_color="E6F3FF", end_color="E6F3FF", fill_type="solid"
        )
        # Define fill for baseline config rows (light green)
        baseline_fill = PatternFill(
            start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"
        )
        baseline_font = Font(bold=True)
        # Use module-level BASELINE_CONFIG

        # Parse baseline config to get expected values for each config column
        # Config format: T{text}_F{feedback}_H{hand}_C{common}_P{prompt}_I{images}_R{remember}_S{full_steps}_E{temp}_M{max}
        baseline_values = {}
        baseline_parts = BASELINE_CONFIG.split("_")
        for part in baseline_parts:
            if part.startswith("T"):
                # text_only: T0 = False, T1 = True
                baseline_values["text_only"] = part[1:] == "1"
            elif part.startswith("F"):
                # feedback_type: Fn = none, Fs = simple, Fd = detailed
                code = part[1:]
                baseline_values["feedback_type"] = {
                    "n": "none",
                    "s": "simple",
                    "d": "detailed",
                }.get(code, code)
            elif part.startswith("H"):
                # hand_transparency: H60 = 60
                baseline_values["hand_transparency"] = int(part[1:])
            elif part.startswith("C"):
                # include_common_sense: C0 = False, C1 = True
                baseline_values["include_common_sense"] = part[1:] == "1"
            elif part.startswith("P"):
                # prompt_version: P2 = v2
                baseline_values["prompt_version"] = f"v{part[1:]}"
            elif part.startswith("I"):
                # previous_image: I0 = none, I1 = color, I2 = grayscale
                baseline_values["previous_image"] = {
                    "0": "none",
                    "1": "color",
                    "2": "grayscale",
                }.get(part[1:], part[1:])
            elif part.startswith("R"):
                # use_memory: R0 = False, R1 = True
                baseline_values["use_memory"] = part[1:] == "1"
            elif part.startswith("S"):
                # full_steps: S0 = False, S1 = True
                baseline_values["full_steps"] = part[1:] == "1"
            elif part.startswith("E"):
                # temperature: E0 = 0.0, E60 = 0.6
                baseline_values["temperature"] = int(part[1:]) / 100.0
            elif part.startswith("M"):
                # max_completion_tokens: M4096 = 4096
                baseline_values["max_completion_tokens"] = int(part[1:])

        # Build mapping of column index to baseline value (for config columns G-N)
        config_col_baseline = {}
        for col_idx, header in enumerate(headers, start=1):
            if header in baseline_values:
                config_col_baseline[col_idx] = (header, baseline_values[header])

        # Define fill for cells that differ from baseline (light yellow)
        differ_from_baseline_fill = PatternFill(
            start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"
        )

        # Define fill and font for ERR cells (white text on red background)
        error_fill = PatternFill(
            start_color="FF0000", end_color="FF0000", fill_type="solid"
        )
        error_font = Font(bold=True, color="FFFFFF")

        # Find column indices for % change calculation
        success_pct_col = None
        pct_change_col = None
        success_stdev_col = None  # Column index for success stdev (1-indexed for Excel)
        config_col = None  # Column index for config (0-indexed in row_data)
        model_col = None  # Column index for model (0-indexed in row_data)
        test_set_col = None  # Column index for test_set_name (0-indexed in row_data)
        for col_idx, header in enumerate(headers, start=1):
            if header == "success %":
                success_pct_col = col_idx
            elif header == "% change":
                pct_change_col = col_idx
            elif header == "success stdev":
                success_stdev_col = col_idx
            elif header == "config":
                config_col = col_idx - 1  # 0-indexed for row_data access
            elif header == "model":
                model_col = col_idx - 1
            elif header == "test_set_name":
                test_set_col = col_idx - 1

        # Helper function to normalize config strings for baseline comparison
        def normalize_config_for_baseline(cfg: str) -> str:
            """Remove M, E, P, C parts from config string for baseline comparison."""
            if not cfg:
                return ""
            parts = cfg.split("_")
            # Keep only parts that don't start with M, E, P, or C
            filtered = [p for p in parts if p and p[0] not in ("M", "E", "P", "C")]
            return "_".join(filtered)

        # Build mapping of (model, test_set_name) -> (baseline success %, baseline stdev)
        # Only consider AVG rows (not individual rep rows) for baseline lookup
        baseline_success_map: dict[tuple, tuple[float, float | None]] = {}
        if (
            success_pct_col
            and config_col is not None
            and model_col is not None
            and test_set_col is not None
        ):
            for row_data in rows:
                if len(row_data) > max(
                    config_col, model_col, test_set_col, success_pct_col - 1
                ):
                    config_val = row_data[config_col]
                    # Check if this is a baseline config AVG row (ignoring M, E, P, C)
                    rep_val = row_data[3] if len(row_data) > 3 else ""
                    if (
                        normalize_config_for_baseline(config_val)
                        == normalize_config_for_baseline(BASELINE_CONFIG)
                        and rep_val == "AVG"
                    ):
                        model_val = row_data[model_col]
                        test_set_val = row_data[test_set_col]
                        success_val = row_data[success_pct_col - 1]  # 0-indexed
                        # Get stdev value if available
                        stdev_val = None
                        if success_stdev_col and len(row_data) > success_stdev_col - 1:
                            stdev_candidate = row_data[success_stdev_col - 1]
                            if isinstance(
                                stdev_candidate, (int, float)
                            ) and not isinstance(stdev_candidate, bool):
                                stdev_val = stdev_candidate
                        if isinstance(success_val, (int, float)) and not isinstance(
                            success_val, bool
                        ):
                            baseline_success_map[(model_val, test_set_val)] = (
                                success_val,
                                stdev_val,
                            )

        # Write data rows
        for row_idx, row_data in enumerate(rows, start=2):
            # Check if this is a rep row (rep column contains "rep1", "rep2", etc.)
            is_rep_row = False
            if rep_col and rep_col - 1 < len(row_data):
                rep_value = row_data[rep_col - 1]
                if isinstance(rep_value, str) and rep_value.startswith("rep"):
                    is_rep_row = True

            # Check if this is a baseline config row
            # Method 1: config column exactly matches BASELINE_CONFIG string (ignoring M, E, P values)
            # Method 2: all individual config values match baseline (for cases where config string differs)
            is_baseline_row = False
            if len(row_data) > 2:  # config is column index 2 (0-based)
                config_value = row_data[2]

                # Compare config strings ignoring M (max_tokens), E (temperature), P (prompt_version), C (include_common_sense)
                if normalize_config_for_baseline(
                    config_value
                ) == normalize_config_for_baseline(BASELINE_CONFIG):
                    is_baseline_row = True
                else:
                    # Check if all config column values match baseline
                    # Ignore max_completion_tokens, temperature, prompt_version, include_common_sense
                    ignored_configs = {
                        "max_completion_tokens",
                        "temperature",
                        "prompt_version",
                        "include_common_sense",
                    }
                    all_match_baseline = True
                    has_any_config = False
                    for col_idx, (
                        header_name,
                        baseline_val,
                    ) in config_col_baseline.items():
                        if header_name in ignored_configs:
                            continue
                        if col_idx - 1 < len(row_data):  # col_idx is 1-based
                            cell_val = row_data[col_idx - 1]
                            if cell_val is not None and cell_val != "":
                                has_any_config = True
                                # Handle string comparisons case-insensitively
                                if isinstance(cell_val, str) and isinstance(
                                    baseline_val, str
                                ):
                                    if cell_val.lower() != baseline_val.lower():
                                        all_match_baseline = False
                                        break
                                else:
                                    if cell_val != baseline_val:
                                        all_match_baseline = False
                                        break
                    if has_any_config and all_match_baseline:
                        is_baseline_row = True

            # Calculate % change from baseline for AVG rows
            pct_change_value = ""
            is_significant_change = False  # Track if ranges don't overlap
            rep_val = row_data[3] if len(row_data) > 3 else ""
            if (
                rep_val == "AVG"
                and not is_baseline_row
                and success_pct_col
                and pct_change_col
                and model_col is not None
                and test_set_col is not None
            ):
                model_val = row_data[model_col] if len(row_data) > model_col else ""
                test_set_val = (
                    row_data[test_set_col] if len(row_data) > test_set_col else ""
                )
                current_success = (
                    row_data[success_pct_col - 1]
                    if len(row_data) > success_pct_col - 1
                    else None
                )
                # Get current row's stdev
                current_stdev = None
                if success_stdev_col and len(row_data) > success_stdev_col - 1:
                    stdev_candidate = row_data[success_stdev_col - 1]
                    if isinstance(stdev_candidate, (int, float)) and not isinstance(
                        stdev_candidate, bool
                    ):
                        current_stdev = stdev_candidate

                baseline_result = baseline_success_map.get((model_val, test_set_val))
                baseline_success = baseline_result[0] if baseline_result else None
                baseline_stdev = baseline_result[1] if baseline_result else None

                if (
                    baseline_success is not None
                    and isinstance(current_success, (int, float))
                    and not isinstance(current_success, bool)
                ):
                    # Calculate percentage point change (e.g., 0.45 -> 0.50 = +5%)
                    pct_change_value = current_success - baseline_success

                    # Check for significance: ranges don't overlap
                    # Range = mean ± stdev
                    # Non-overlap if: baseline_upper < current_lower OR current_upper < baseline_lower
                    if baseline_stdev is not None and current_stdev is not None:
                        baseline_lower = baseline_success - baseline_stdev
                        baseline_upper = baseline_success + baseline_stdev
                        current_lower = current_success - current_stdev
                        current_upper = current_success + current_stdev

                        # Significant if ranges don't overlap
                        if (
                            baseline_upper < current_lower
                            or current_upper < baseline_lower
                        ):
                            is_significant_change = True
                    pct_change_value = current_success - baseline_success

            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                # Apply bold right border for column group separators
                if col_idx in bold_border_col_indices:
                    cell.border = bold_right_border
                else:
                    cell.border = thin_border

                # Apply baseline styling (light green background, bold) - highest priority
                if is_baseline_row:
                    cell.fill = baseline_fill
                    cell.font = baseline_font
                # Apply light blue background for rep rows
                elif is_rep_row:
                    cell.fill = light_blue_fill
                # Highlight config columns that differ from baseline (not for baseline or rep rows)
                elif col_idx in config_col_baseline:
                    header_name, baseline_val = config_col_baseline[col_idx]
                    # Normalize values for comparison
                    cell_val = value
                    # Handle string comparisons case-insensitively
                    if isinstance(cell_val, str) and isinstance(baseline_val, str):
                        differs = cell_val.lower() != baseline_val.lower()
                    else:
                        differs = cell_val != baseline_val
                    if differs:
                        cell.fill = differ_from_baseline_fill

                # Special handling for ERR values (inconsistent config across reps)
                if value == "ERR":
                    cell.font = error_font
                    cell.fill = error_fill
                    cell.alignment = Alignment(horizontal="center")
                    continue

                # Special handling for % change column
                if col_idx == pct_change_col:
                    if isinstance(pct_change_value, (int, float)):
                        cell.value = pct_change_value
                        cell.number_format = "+0.0%;-0.0%;0.0%"
                        cell.alignment = Alignment(horizontal="center")
                        # Color based on positive/negative change
                        if pct_change_value > 0.001:  # Positive change (green)
                            cell.font = Font(color="008000")
                        elif pct_change_value < -0.001:  # Negative change (red)
                            cell.font = Font(color="FF0000")
                        # Highlight yellow if ranges don't overlap (significant change)
                        if is_significant_change:
                            cell.fill = PatternFill(
                                start_color="FFFF00",
                                end_color="FFFF00",
                                fill_type="solid",
                            )
                    elif is_baseline_row:
                        cell.value = "—"  # Em dash for baseline row
                        cell.alignment = Alignment(horizontal="center")
                    continue

                # Special handling for test_completed column
                if col_idx == test_completed_col:
                    if value is True:
                        cell.value = "✓"
                        cell.font = green_font
                        cell.alignment = Alignment(horizontal="center")
                    elif isinstance(value, (int, float)):
                        # Show as percentage with light red background
                        cell.number_format = "0%"
                        cell.alignment = Alignment(horizontal="center")
                        cell.fill = light_red_fill
                    continue

                # Special handling for hand_transparency column
                if col_idx == hand_transparency_col:
                    if value == 60:
                        cell.value = "✓"
                        cell.font = green_font
                        cell.alignment = Alignment(horizontal="center")
                    elif value == 0:
                        cell.value = "✗"
                        cell.font = red_font
                        cell.alignment = Alignment(horizontal="center")
                    else:
                        cell.alignment = Alignment(horizontal="center")
                    continue

                # Special handling for feedback_type column
                if col_idx == feedback_type_col:
                    if isinstance(value, str):
                        if value.lower() == "simple":
                            cell.value = "±"
                            cell.font = Font(bold=True, color="FFA500")  # Orange
                            cell.alignment = Alignment(horizontal="center")
                        elif value.lower() == "detailed":
                            cell.value = "☰"
                            cell.font = Font(bold=True, color="008000")  # Green
                            cell.alignment = Alignment(horizontal="center")
                        elif value.lower() == "none":
                            cell.value = "✗"
                            cell.font = red_font
                            cell.alignment = Alignment(horizontal="center")
                        else:
                            cell.alignment = Alignment(horizontal="center")
                    continue

                # Special handling for previous_image column
                if col_idx == previous_image_col:
                    if isinstance(value, str):
                        if value.lower() == "color":
                            cell.value = "■"  # Black square for color
                            cell.font = Font(bold=True, color="000000")  # Black
                            cell.alignment = Alignment(horizontal="center")
                        elif value.lower() == "grayscale":
                            cell.value = "□"  # White square for grayscale
                            cell.font = Font(bold=True, color="808080")  # Gray
                            cell.alignment = Alignment(horizontal="center")
                        elif value.lower() == "none":
                            cell.value = "✗"
                            cell.font = red_font
                            cell.alignment = Alignment(horizontal="center")
                        else:
                            cell.alignment = Alignment(horizontal="center")
                    continue

                # Format percentage columns
                if col_idx in percentage_cols and isinstance(value, (int, float)):
                    cell.number_format = "0.0%"
                    cell.alignment = Alignment(horizontal="right")
                    # Dark green bold for success_percentage column
                    if col_idx == success_percentage_col:
                        cell.font = success_font
                    # Dark red bold for fail_percentage column
                    elif col_idx == fail_percentage_col:
                        cell.font = fail_font
                    # Red background for fail api % column if non-zero
                    elif col_idx == fail_api_percentage_col:
                        if value > 0:
                            cell.fill = light_red_fill
                elif value is None:
                    cell.value = "✗"
                    cell.font = red_font
                    cell.alignment = Alignment(horizontal="center")
                elif isinstance(value, bool):
                    if value:
                        cell.value = "✓"
                        cell.font = green_font
                    else:
                        cell.value = "✗"
                        cell.font = red_font
                    cell.alignment = Alignment(horizontal="center")
                elif isinstance(value, str) and value.lower() in ("yes", "no", "none"):
                    if value.lower() == "yes":
                        cell.value = "✓"
                        cell.font = green_font
                    else:
                        cell.value = "✗"
                        cell.font = red_font
                    cell.alignment = Alignment(horizontal="center")
                elif isinstance(value, (int, float)):
                    # Integer count columns should display as rounded integers
                    if col_idx in integer_count_cols:
                        cell.value = round(value)
                        cell.number_format = "0"
                    elif isinstance(value, float):
                        cell.number_format = "0.000"
                    cell.alignment = Alignment(horizontal="right")

        # Auto-adjust column widths based on original data (before cell transformations)
        for col_idx in range(len(headers)):
            column_letter = get_column_letter(col_idx + 1)
            col_idx_1based = col_idx + 1  # Convert to 1-based for percentage_cols check

            # Icon columns get fixed narrow width
            if col_idx_1based in icon_cols:
                ws.column_dimensions[column_letter].width = 5.5
                continue

            # Calculate max width based on original row data
            max_length = 0
            for row_data in rows:
                if col_idx < len(row_data):
                    value = row_data[col_idx]
                    if value is not None:
                        # For percentage columns, estimate formatted width (e.g., "85.5%")
                        if col_idx_1based in percentage_cols and isinstance(
                            value, (int, float)
                        ):
                            formatted = f"{value * 100:.1f}%"
                            max_length = max(max_length, len(formatted))
                        # For booleans and yes/no/none strings, they become single char symbols
                        elif isinstance(value, bool):
                            max_length = max(max_length, 1)
                        elif isinstance(value, str) and value.lower() in (
                            "yes",
                            "no",
                            "none",
                        ):
                            max_length = max(max_length, 1)
                        # For floats, estimate formatted width (e.g., "0.123")
                        elif isinstance(value, float):
                            formatted = f"{value:.3f}"
                            max_length = max(max_length, len(formatted))
                        else:
                            max_length = max(max_length, len(str(value)))
            # Minimal padding, just enough to fit content
            adjusted_width = max(max_length + 1, 5.5)  # Minimum width of 5.5
            ws.column_dimensions[column_letter].width = adjusted_width

        # Hide the model_path column
        if model_path_col:
            ws.column_dimensions[get_column_letter(model_path_col)].hidden = True

        # Set header row height to accommodate rotated text
        # Calculate based on longest header (with added spacing and underscore replacement)
        max_header_len = max(len("        " + h.replace("_", " ")) for h in headers)
        # Approximate: each character needs about 7 points of height when rotated
        header_row_height = max(max_header_len * 7, 100)
        ws.row_dimensions[1].height = header_row_height

        # Freeze the header row
        ws.freeze_panes = "A2"

        # Add autofilter
        ws.auto_filter.ref = ws.dimensions

        return wb

    def _parse_rep_suffix(self, model_dir: str) -> Tuple[str, str | None]:
        """Parse a model directory name to extract base name and rep suffix.

        Args:
            model_dir: Model directory name, e.g.:
                       "gpt-4o--T0_Fd--rep1"
                       "gpt-4o--T0_Fd--rep2_T0_Fd..."

        Returns:
            Tuple of (base_name, rep_suffix) where rep_suffix is None if no rep pattern found.
            e.g. ("gpt-4o--T0_Fd", "rep1") or ("gpt-4o--T0_Fd", None)
        """
        import re

        # Match --repN at end OR --repN_ followed by more text
        # Capture everything before --repN as base_name
        match = re.search(r"--(rep\d+)(?:_.*)?$", model_dir)
        if match:
            # Base name is everything before the --repN
            base_name = model_dir[: match.start()]
            return base_name, match.group(1)
        return model_dir, None

    def _group_model_dirs_by_base(
        self, model_dirs: List[str]
    ) -> dict[str, List[Tuple[str, str | None]]]:
        """Group model directories by their base name (without rep suffix).

        Returns:
            Dict mapping base_name -> list of (full_model_dir, rep_suffix) tuples
        """
        from collections import defaultdict

        groups: dict[str, List[Tuple[str, str | None]]] = defaultdict(list)
        for model_dir in model_dirs:
            base_name, rep_suffix = self._parse_rep_suffix(model_dir)
            groups[base_name].append((model_dir, rep_suffix))
        return groups

    def _config_diagonal_sort_key(self, base_model_name: str) -> tuple:
        """Generate a sort key for config strings to create diagonal pattern.

        For each model, sorts config variations so that the "varying" config
        (differing from baseline) forms a diagonal pattern in the spreadsheet.

        The diagonal order:
        1. feedback_type detailed (Fd)
        2. feedback_type none (Fn)
        3. full_steps (S0)
        4. hand_transparency/show_hand (H0)
        5. previous_image (I0 or I2)
        6. text_only (T1)
        7. use_memory (R0)
        8. baseline (nothing differs) - very bottom

        Args:
            base_model_name: The model directory name (may contain --config_string)

        Returns:
            A tuple for sorting that groups by model and orders configs for diagonal
        """
        # Extract model part and config part
        if "--" in base_model_name:
            model_part, config_part = base_model_name.split("--", 1)
        else:
            model_part = base_model_name
            config_part = ""

        # Parse baseline config values (from BASELINE_CONFIG = "T0_Fs_H60_C0_P2_I1_R1_S1_E0_M4096")
        baseline = {
            "T": "0",  # text_only: 0=False
            "F": "s",  # feedback_type: s=simple
            "H": "60",  # hand_transparency: 60
            "I": "1",  # previous_image: 1=color
            "R": "1",  # use_memory: 1=True
            "S": "1",  # full_steps: 1=True
        }

        # Parse config string to extract values
        config_values = {}
        if config_part:
            parts = config_part.split("_")
            for part in parts:
                if part and len(part) >= 2:
                    prefix = part[0]
                    value = part[1:]
                    config_values[prefix] = value

        # Determine diagonal position based on which parameter differs from baseline
        # Order: Fd, Fn, S0, H0, I0/I2, T1, R0, baseline
        diagonal_position = 100  # Default: baseline goes last

        # Check each config in order
        f_val = config_values.get("F")
        if f_val == "d":  # feedback_type detailed
            diagonal_position = 0
        elif f_val == "n":  # feedback_type none
            diagonal_position = 1
        elif config_values.get("S") != baseline.get("S"):  # full_steps differs
            diagonal_position = 2
        elif config_values.get("H") != baseline.get("H"):  # hand_transparency differs
            diagonal_position = 3
        elif config_values.get("I") != baseline.get("I"):  # previous_image differs
            diagonal_position = 4
        elif config_values.get("T") != baseline.get("T"):  # text_only differs
            diagonal_position = 5
        elif config_values.get("R") != baseline.get("R"):  # use_memory differs
            diagonal_position = 6

        return (model_part, diagonal_position, config_part)

    def _average_rows(self, rows: List[List[Any]], config_keys: List[str]) -> List[Any]:
        """Compute an average row from multiple result rows.

        Numeric values are averaged, non-numeric values use the first row's value.
        Boolean values should be identical across all rows (config settings).
        String values should be identical across all rows (config settings).
        """
        if not rows:
            return []
        if len(rows) == 1:
            return rows[0].copy()

        avg_row = []
        for col_idx in range(len(rows[0])):
            values = [row[col_idx] for row in rows if row[col_idx] is not None]
            if not values:
                avg_row.append(None)
            # Check for booleans first (before int/float since bool is subclass of int)
            elif all(isinstance(v, bool) for v in values):
                # All boolean values should be the same (config setting)
                if len(set(values)) > 1:
                    avg_row.append("ERR")
                else:
                    avg_row.append(rows[0][col_idx])
            elif all(isinstance(v, str) for v in values):
                # All string values should be the same (config setting)
                if len(set(values)) > 1:
                    avg_row.append("ERR")
                else:
                    avg_row.append(rows[0][col_idx])
            elif all(
                isinstance(v, (int, float)) and not isinstance(v, bool) for v in values
            ):
                # Average numeric values (excluding booleans)
                avg_row.append(sum(values) / len(values))
            else:
                # Use first row's value for mixed types
                avg_row.append(rows[0][col_idx])

        # Calculate success range and stdev for AVG+ALL rows only
        # Check if all rows being averaged have test_set_name == "ALL" (index 5)
        is_all_rows = all(len(row) > 5 and row[5] == "ALL" for row in rows)

        if is_all_rows:
            import statistics

            headers = self.get_headers(config_keys)
            success_pct_idx = (
                headers.index("success %") if "success %" in headers else None
            )
            range_idx = (
                headers.index("success range") if "success range" in headers else None
            )
            stdev_idx = (
                headers.index("success stdev") if "success stdev" in headers else None
            )

            if (
                success_pct_idx is not None
                and range_idx is not None
                and stdev_idx is not None
            ):
                # Collect success % values from all rows
                success_values = [
                    row[success_pct_idx]
                    for row in rows
                    if isinstance(row[success_pct_idx], (int, float))
                    and not isinstance(row[success_pct_idx], bool)
                ]
                if success_values:
                    min_val = min(success_values)
                    max_val = max(success_values)
                    # Format as percentage strings (values are 0-1)
                    avg_row[range_idx] = f"{min_val*100:.1f}% - {max_val*100:.1f}%"
                    # Calculate stdev (as decimal 0-1 to match other percentage columns)
                    if len(success_values) > 1:
                        avg_row[stdev_idx] = statistics.stdev(success_values)
                    else:
                        avg_row[stdev_idx] = 0.0

        return avg_row

    def print_all_results(
        self,
        model_dirs: List[str] | None = None,
        test_list: List[str] | None = None,
        output_file: str | None = None,
    ) -> Tuple[Workbook, str] | None:
        """Print aggregated results for all models and tests.

        Args:
            model_dirs: List of model directory names. If None, auto-discovers from directory.
            test_list: List of test names. If None, auto-discovers from directory.
            output_file: Path to output CSV file. Defaults to {test_dir}/results.csv.

        Returns:
            Tuple of (workbook, xlsx_file_path) or None if no results.
        """
        # Auto-discover if not provided
        if model_dirs is None or test_list is None:
            discovered_model_dirs, discovered_tests = self.discover_tests_and_models()
            if model_dirs is None:
                model_dirs = discovered_model_dirs
            if test_list is None:
                test_list = discovered_tests
            print(
                f"Auto-discovered {len(model_dirs)} model directories and {len(test_list)} tests under {self.test_dir}."
            )

        if not model_dirs or not test_list:
            print("No models or tests found to process.")
            return None

        # Ensure the test directory exists
        if not os.path.exists(self.test_dir):
            os.makedirs(self.test_dir)

        # Collect all config keys
        config_keys = self.collect_config_keys(test_list)

        # Determine output files
        csv_file = output_file or f"{self.test_dir}/results.csv"
        xlsx_file = (
            csv_file.replace(".csv", ".xlsx")
            if csv_file.endswith(".csv")
            else f"{csv_file}.xlsx"
        )

        # Test if we can write to the output location, fallback to local if not
        try:
            with open(csv_file, "w", encoding="utf-8") as f:
                pass  # Just test if we can open for writing
        except OSError as e:
            # Fallback to local Generated directory
            local_csv = f"Generated/results.csv"
            local_xlsx = f"Generated/results.xlsx"
            print(
                f"Warning: Cannot write to {csv_file} ({e}). Falling back to {local_csv}"
            )
            os.makedirs("Generated", exist_ok=True)
            csv_file = local_csv
            xlsx_file = local_xlsx

        # Group model directories by base name to handle --repN suffixes
        model_groups = self._group_model_dirs_by_base(model_dirs)

        # Collect all results rows
        all_rows: List[List[Any]] = []

        # Write CSV and collect rows for Excel
        with open(csv_file, "w", encoding="utf-8") as file:
            ResultStat.add_header_to_csv(file, config_keys)

            # Process each base model group (sorted to create diagonal pattern)
            for base_model_name in sorted(
                model_groups.keys(), key=self._config_diagonal_sort_key
            ):
                group = model_groups[base_model_name]
                has_reps = any(rep_suffix is not None for _, rep_suffix in group)

                for test in test_list:
                    if has_reps:
                        # Has rep suffix(es): collect rows for averaging
                        rep_rows = []
                        for model_dir, rep_suffix in sorted(
                            group, key=lambda x: x[1] or ""
                        ):
                            combined_results, combined_config = (
                                self.print_results_for_model(
                                    [test], model_dir, file, config_keys=config_keys
                                )
                            )
                            row = self.collect_results_row(
                                combined_results, combined_config, config_keys
                            )
                            if row:
                                # Modify row: use base model name, set rep column
                                # Split base_model_name into model and config
                                if "--" in base_model_name:
                                    model_part, config_part = base_model_name.split(
                                        "--", 1
                                    )
                                else:
                                    model_part, config_part = base_model_name, ""
                                row[1] = model_part  # model column
                                row[2] = config_part  # config column
                                row[3] = rep_suffix or ""  # rep column
                                row[4] = (
                                    "-"  # rep_count column (dash for individual reps)
                                )
                                all_rows.append(row)
                                rep_rows.append(row.copy())  # Copy for averaging

                        # Add average row (even for single rep)
                        if rep_rows:
                            avg_row = self._average_rows(rep_rows, config_keys)
                            # Split base_model_name into model and config
                            if "--" in base_model_name:
                                model_part, config_part = base_model_name.split("--", 1)
                            else:
                                model_part, config_part = base_model_name, ""
                            avg_row[1] = model_part  # model
                            avg_row[2] = config_part  # config
                            avg_row[3] = "AVG"  # rep column
                            avg_row[4] = len(
                                rep_rows
                            )  # rep_count column (number of reps averaged)
                            # test_completed: True if all complete, else average of percentages
                            completions = [r[0] for r in rep_rows]
                            if all(c is True for c in completions):
                                avg_row[0] = True
                            else:
                                # Convert True to 1.0 for averaging
                                numeric_completions = [
                                    1.0 if c is True else c for c in completions
                                ]
                                avg_row[0] = sum(numeric_completions) / len(
                                    numeric_completions
                                )
                            all_rows.append(avg_row)
                    else:
                        # Single model or no rep suffix: process normally
                        model_dir, rep_suffix = group[0]
                        combined_results, combined_config = (
                            self.print_results_for_model(
                                [test], model_dir, file, config_keys=config_keys
                            )
                        )
                        row = self.collect_results_row(
                            combined_results, combined_config, config_keys
                        )
                        if row:
                            # Still set base model name and rep for single-rep cases
                            # Split base_model_name into model and config
                            if "--" in base_model_name:
                                model_part, config_part = base_model_name.split("--", 1)
                            else:
                                model_part, config_part = base_model_name, ""
                            row[1] = model_part
                            row[2] = config_part
                            row[3] = rep_suffix or ""
                            row[4] = "-"  # rep_count column
                            all_rows.append(row)

                # Print combined data for all discovered tests
                if has_reps:
                    rep_rows = []
                    for model_dir, rep_suffix in sorted(
                        group, key=lambda x: x[1] or ""
                    ):
                        combined_results, combined_config = (
                            self.print_results_for_model(
                                test_list, model_dir, file, config_keys=config_keys
                            )
                        )
                        row = self.collect_results_row(
                            combined_results, combined_config, config_keys
                        )
                        if row:
                            # Split base_model_name into model and config
                            if "--" in base_model_name:
                                model_part, config_part = base_model_name.split("--", 1)
                            else:
                                model_part, config_part = base_model_name, ""
                            row[1] = model_part
                            row[2] = config_part
                            row[3] = rep_suffix or ""
                            row[4] = "-"  # rep_count column
                            all_rows.append(row)
                            rep_rows.append(row.copy())  # Copy for averaging

                    # Add average row (even for single rep)
                    if rep_rows:
                        avg_row = self._average_rows(rep_rows, config_keys)
                        # Split base_model_name into model and config
                        if "--" in base_model_name:
                            model_part, config_part = base_model_name.split("--", 1)
                        else:
                            model_part, config_part = base_model_name, ""
                        avg_row[1] = model_part
                        avg_row[2] = config_part
                        avg_row[3] = "AVG"
                        avg_row[4] = len(rep_rows)  # rep_count column
                        # test_completed: True if all complete, else average of percentages
                        completions = [r[0] for r in rep_rows]
                        if all(c is True for c in completions):
                            avg_row[0] = True
                        else:
                            # Convert True to 1.0 for averaging
                            numeric_completions = [
                                1.0 if c is True else c for c in completions
                            ]
                            avg_row[0] = sum(numeric_completions) / len(
                                numeric_completions
                            )
                        all_rows.append(avg_row)
                else:
                    model_dir, rep_suffix = group[0]
                    combined_results, combined_config = self.print_results_for_model(
                        test_list, model_dir, file, config_keys=config_keys
                    )
                    row = self.collect_results_row(
                        combined_results, combined_config, config_keys
                    )
                    if row:
                        # Split base_model_name into model and config
                        if "--" in base_model_name:
                            model_part, config_part = base_model_name.split("--", 1)
                        else:
                            model_part, config_part = base_model_name, ""
                        row[1] = model_part
                        row[2] = config_part
                        row[3] = rep_suffix or ""
                        row[4] = "-"  # rep_count column
                        all_rows.append(row)

        print(f"--= Wrote CSV results to {csv_file}")

        # Filter rows to only include 100% complete tests
        # test_completed is the first column (index 0), True means 100% complete
        complete_rows = [row for row in all_rows if row[0] is True]
        print(
            f"--= Filtered to {len(complete_rows)} complete rows (out of {len(all_rows)} total)"
        )

        # Write Excel file with filtered results first (creates workbook)
        headers = self.get_headers(config_keys)
        wb = self.write_xlsx(
            xlsx_file, headers, complete_rows, config_keys, sheet_name="Results"
        )

        # Add "Results All" sheet with all rows (including incomplete)
        wb = self.write_xlsx(
            xlsx_file,
            headers,
            all_rows,
            config_keys,
            workbook=wb,
            sheet_name="Results All",
        )

        return wb, xlsx_file

    def generate_plan_performance_report(
        self,
        output_file: str | None = None,
        workbook: Workbook | None = None,
    ) -> Workbook | None:
        """Generate a report analyzing performance across all plans.

        This report reads from test_results.json files (which contain orig_step_count)
        to aggregate statistics for each unique plan across all test/model combinations.
        Plans are identified by task_name, with failed plans indicated by task_failed=true.

        Args:
            output_file: Path to output CSV file. Defaults to {test_dir}/plan_performance.csv.
            workbook: Optional existing workbook to add sheet to.

        Returns:
            The workbook with plan performance sheet added, or None if no plans found.
        """
        from collections import defaultdict
        from dataclasses import dataclass, field

        @dataclass
        class PlanStats:
            """Statistics for a single plan across all test runs."""

            plan_name: str
            total_runs: int = 0
            successes: int = 0
            failures: int = 0
            success_steps: list = field(default_factory=list)
            failure_steps: list = field(default_factory=list)
            error_counts: dict = field(default_factory=dict)  # error_msg -> count
            # Track placeholder values: error_msg -> {placeholder -> {object_type -> count}}
            # e.g., "slice : {X} is sliced" -> {"X": {"Bread": 2, "Egg": 1}, "Y": {...}}
            placeholder_values: dict = field(default_factory=dict)

        # Dictionary to collect stats for each unique plan
        plan_stats: dict[str, PlanStats] = defaultdict(lambda: PlanStats(plan_name=""))

        if not os.path.exists(self.test_dir):
            print(f"Test directory not found: {self.test_dir}")
            return

        # First, count tests and models for progress reporting
        test_names = [
            t
            for t in os.listdir(self.test_dir)
            if os.path.isdir(os.path.join(self.test_dir, t))
        ]
        total_tests = len(test_names)
        print(f"Scanning {total_tests} test directories for plan performance...")

        plans_processed = 0

        # Scan all test directories
        for test_idx, test_name in enumerate(test_names, start=1):
            test_path = os.path.join(self.test_dir, test_name)
            if not os.path.isdir(test_path):
                continue

            # Count models in this test
            model_dirs_list = [
                m
                for m in os.listdir(test_path)
                if os.path.isdir(os.path.join(test_path, m))
            ]
            print(
                f"  [{test_idx}/{total_tests}] {test_name}: {len(model_dirs_list)} models"
            )

            # Scan all model directories within each test
            for model_idx, model_dir in enumerate(model_dirs_list, start=1):
                model_path = os.path.join(test_path, model_dir)
                if not os.path.isdir(model_path):
                    continue

                # Read test_results.json instead of individual plan.json files
                test_results_path = os.path.join(model_path, "test_results.json")
                if not os.path.exists(test_results_path):
                    continue

                test_results_data = self._read_json_cached(test_results_path)

                results_list = test_results_data.get("test_results", [])
                model_plans = 0

                for result in results_list:
                    task_name = result.get("task_name", "")
                    if not task_name:
                        continue

                    task_failed = result.get("task_failed", False)
                    # test_step_count is the number of steps taken during the test
                    num_steps = result.get("test_step_count", 0)

                    # Get or create stats for this plan
                    if task_name not in plan_stats:
                        plan_stats[task_name] = PlanStats(plan_name=task_name)

                    stats = plan_stats[task_name]
                    stats.total_runs += 1

                    if task_failed:
                        stats.failures += 1
                        if num_steps > 0:
                            stats.failure_steps.append(num_steps)
                    else:
                        stats.successes += 1
                        if num_steps > 0:
                            stats.success_steps.append(num_steps)

                    # Collect error messages from step_errors
                    step_errors = result.get("step_errors", [])
                    if step_errors:
                        import re

                        # Words to exclude from object matching (common English words)
                        excluded_words = {
                            "Cannot",
                            "The",
                            "This",
                            "That",
                            "There",
                            "Here",
                            "What",
                            "When",
                            "Where",
                            "Which",
                            "While",
                            "With",
                            "Without",
                            "Could",
                            "Would",
                            "Should",
                            "Must",
                            "Will",
                            "Shall",
                            "Have",
                            "Has",
                            "Had",
                            "Does",
                            "Did",
                            "Done",
                            "Been",
                            "Being",
                            "Was",
                            "Were",
                            "Are",
                            "Is",
                            "Am",
                            "Be",
                            "Not",
                            "But",
                            "And",
                            "For",
                            "From",
                            "Into",
                            "Onto",
                            "Over",
                            "Under",
                            "Above",
                            "Below",
                            "Between",
                            "Among",
                            "Through",
                            "During",
                            "Before",
                            "After",
                            "Since",
                            "Until",
                            "Already",
                            "Also",
                            "Always",
                            "Never",
                            "Ever",
                            "Only",
                            "Just",
                            "Still",
                            "Even",
                            "Very",
                            "Too",
                            "Much",
                            "Many",
                            "Some",
                            "Any",
                            "All",
                            "Each",
                            "Every",
                            "Both",
                            "Either",
                            "Neither",
                            "Other",
                            "Another",
                            "Such",
                            "Same",
                            "Different",
                            "PUT",
                            "GET",
                            "Ran",
                            "Failed",
                            "Error",
                            "Invalid",
                            "Missing",
                            "Expected",
                            "Unexpected",
                            "Unable",
                            "Slice",
                            "Sliced",
                            "Types",
                            "Type",
                            "Action",
                            "Object",
                            "Objects",
                            "Specifier",
                            "Candidate",
                            "Candidates",
                            "None",
                            "True",
                            "False",
                        }
                        # Pattern to match object names with suffixes
                        object_with_suffix_pattern = re.compile(
                            r"[A-Z][a-zA-Z]*(?:_[a-zA-Z0-9]+)+(?:\(Clone\))?"
                        )
                        # Pattern to match simple CamelCase object names (e.g., WineBottle, SinkBasin)
                        simple_object_pattern = re.compile(
                            r"\b([A-Z][a-z]+(?:[A-Z][a-z]+)+)\b"
                        )

                        def extract_object_type(obj_name: str) -> str:
                            """Extract the base object type from a full object name.
                            E.g., 'Bread_fe4bb3e3' -> 'Bread', 'Egg_Cracked_20(Clone)' -> 'Egg_Cracked'
                            """
                            # Remove (Clone) suffix if present
                            name = obj_name.replace("(Clone)", "")
                            # Split by underscore and take parts that are not hex/numeric
                            parts = name.split("_")
                            type_parts = []
                            for part in parts:
                                # Stop if we hit a hex-like or pure numeric suffix
                                if re.match(r"^[a-f0-9]{6,}$", part, re.IGNORECASE):
                                    break
                                if re.match(r"^\d+$", part):
                                    break
                                type_parts.append(part)
                            return "_".join(type_parts) if type_parts else parts[0]

                        def find_objects_in_message(msg: str) -> list:
                            """Find all object references in an error message."""
                            result = []
                            # First find objects with suffixes (e.g., Bread_fe4bb3e3)
                            for match in object_with_suffix_pattern.findall(msg):
                                if match not in result:
                                    result.append(match)
                            # Find simple CamelCase objects (e.g., WineBottle, SinkBasin)
                            for match in simple_object_pattern.findall(msg):
                                if match not in excluded_words and match not in result:
                                    # Check it's not already covered by a suffixed version
                                    if not any(
                                        m.startswith(match + "_") for m in result
                                    ):
                                        result.append(match)
                            # Also find single capitalized words that might be objects
                            # (e.g., Pencil, Apple, Bread)
                            single_word_pattern = re.compile(r"\b([A-Z][a-z]{2,})\b")
                            for match in single_word_pattern.findall(msg):
                                if match not in excluded_words and match not in result:
                                    if not any(
                                        m.startswith(match + "_") for m in result
                                    ):
                                        result.append(match)
                            return result

                        for error in step_errors:
                            action_name = error.get("action_name", "")
                            error_msg = error.get("error_msg", "")
                            if error_msg:
                                # Find all unique object names in the error message
                                matches = find_objects_in_message(error_msg)
                                unique_objects = []
                                for m in matches:
                                    if m not in unique_objects:
                                        unique_objects.append(m)

                                # Replace each unique object with {X}, {Y}, {Z}, etc.
                                placeholder_names = ["X", "Y", "Z", "W", "V"]
                                normalized_msg = error_msg
                                object_types = {}  # placeholder_name -> object_type
                                for i, obj_name in enumerate(unique_objects):
                                    ph_name = (
                                        placeholder_names[i]
                                        if i < len(placeholder_names)
                                        else f"O{i}"
                                    )
                                    placeholder = "{" + ph_name + "}"
                                    normalized_msg = normalized_msg.replace(
                                        obj_name, placeholder
                                    )
                                    object_types[ph_name] = extract_object_type(
                                        obj_name
                                    )

                                # Format: "action_name : normalized_error_msg"
                                full_error = (
                                    f"{action_name} : {normalized_msg}"
                                    if action_name
                                    else normalized_msg
                                )
                                stats.error_counts[full_error] = (
                                    stats.error_counts.get(full_error, 0) + 1
                                )

                                # Track placeholder values
                                if full_error not in stats.placeholder_values:
                                    stats.placeholder_values[full_error] = {}
                                for ph_name, obj_type in object_types.items():
                                    if (
                                        ph_name
                                        not in stats.placeholder_values[full_error]
                                    ):
                                        stats.placeholder_values[full_error][
                                            ph_name
                                        ] = {}
                                    stats.placeholder_values[full_error][ph_name][
                                        obj_type
                                    ] = (
                                        stats.placeholder_values[full_error][
                                            ph_name
                                        ].get(obj_type, 0)
                                        + 1
                                    )

                    model_plans += 1
                    plans_processed += 1

                if model_plans > 0:
                    print(
                        f"    [{model_idx}/{len(model_dirs_list)}] {model_dir}: {model_plans} plans"
                    )

        print(f"Processed {plans_processed} total plan runs")

        if not plan_stats:
            print("No plans found to analyze.")
            return

        # Build output data
        def format_placeholder_values(ph_dict: dict) -> str:
            """Format placeholder values as 'Type1 (50%), Type2 (30%)' etc."""
            if not ph_dict:
                return ""
            total = sum(ph_dict.values())
            if total == 0:
                return ""
            # Sort by count descending
            sorted_items = sorted(ph_dict.items(), key=lambda x: x[1], reverse=True)
            parts = [
                f"{obj_type} ({count/total*100:.0f}%)"
                for obj_type, count in sorted_items
            ]
            return ", ".join(parts)

        headers = [
            "plan_name",
            "success_rate",
            "total_runs",
            "successes",
            "failures",
            "success_steps_min",
            "success_steps_max",
            "success_steps_avg",
            "failure_steps_min",
            "failure_steps_max",
            "failure_steps_avg",
            "err1_%",
            "err1_action",
            "err1_message",
            "err1_X",
            "err1_Y",
            "err2_%",
            "err2_action",
            "err2_message",
            "err2_X",
            "err2_Y",
            "err3_%",
            "err3_action",
            "err3_message",
            "err3_X",
            "err3_Y",
        ]

        rows: List[List[Any]] = []
        for plan_name in sorted(plan_stats.keys()):
            stats = plan_stats[plan_name]

            # Calculate success rate
            success_rate = (
                stats.successes / stats.total_runs if stats.total_runs > 0 else 0
            )

            # Calculate success step statistics
            if stats.success_steps:
                success_min = min(stats.success_steps)
                success_max = max(stats.success_steps)
                success_avg = sum(stats.success_steps) / len(stats.success_steps)
            else:
                success_min = success_max = success_avg = None

            # Calculate failure step statistics
            if stats.failure_steps:
                failure_min = min(stats.failure_steps)
                failure_max = max(stats.failure_steps)
                failure_avg = sum(stats.failure_steps) / len(stats.failure_steps)
            else:
                failure_min = failure_max = failure_avg = None

            # Get top 3 most common errors
            sorted_errors = sorted(
                stats.error_counts.items(), key=lambda x: x[1], reverse=True
            )
            total_errors = sum(stats.error_counts.values())

            # Build top 3 errors as (pct, action, message, x_values, y_values) tuples
            top_errors = []
            for i in range(3):
                if i < len(sorted_errors):
                    full_error, count = sorted_errors[i]
                    pct = (count / total_errors * 100) if total_errors > 0 else 0
                    # Parse "action : message" format
                    if " : " in full_error:
                        action, message = full_error.split(" : ", 1)
                    else:
                        action = ""
                        message = full_error
                    # Get placeholder values for this error
                    ph_values = stats.placeholder_values.get(full_error, {})
                    x_values = format_placeholder_values(ph_values.get("X", {}))
                    y_values = format_placeholder_values(ph_values.get("Y", {}))
                    top_errors.append(
                        (f"{pct:.0f}%", action, message, x_values, y_values)
                    )
                else:
                    top_errors.append((None, None, None, None, None))

            rows.append(
                [
                    plan_name,
                    success_rate,
                    stats.total_runs,
                    stats.successes,
                    stats.failures,
                    success_min,
                    success_max,
                    success_avg,
                    failure_min,
                    failure_max,
                    failure_avg,
                    top_errors[0][0],  # err1_%
                    top_errors[0][1],  # err1_action
                    top_errors[0][2],  # err1_message
                    top_errors[0][3],  # err1_X
                    top_errors[0][4],  # err1_Y
                    top_errors[1][0],  # err2_%
                    top_errors[1][1],  # err2_action
                    top_errors[1][2],  # err2_message
                    top_errors[1][3],  # err2_X
                    top_errors[1][4],  # err2_Y
                    top_errors[2][0],  # err3_%
                    top_errors[2][1],  # err3_action
                    top_errors[2][2],  # err3_message
                    top_errors[2][3],  # err3_X
                    top_errors[2][4],  # err3_Y
                ]
            )

        # Sort by success_rate (column index 1) descending
        rows.sort(key=lambda r: r[1] if r[1] is not None else 0, reverse=True)

        # Determine output files
        csv_file = output_file or f"{self.test_dir}/plan_performance.csv"
        xlsx_file = (
            csv_file.replace(".csv", ".xlsx")
            if csv_file.endswith(".csv")
            else f"{csv_file}.xlsx"
        )

        # Write CSV
        with open(csv_file, "w", encoding="utf-8") as f:
            f.write(",".join(headers) + "\n")
            for row in rows:
                csv_row = []
                for val in row:
                    if val is None:
                        csv_row.append("")
                    elif isinstance(val, float):
                        csv_row.append(f"{val:.3f}")
                    else:
                        # Escape commas in plan names
                        str_val = str(val)
                        if "," in str_val:
                            str_val = f'"{str_val}"'
                        csv_row.append(str_val)
                f.write(",".join(csv_row) + "\n")

        print(f"--= Wrote plan performance CSV to {csv_file}")

        # Write Excel file with formatting (add to existing workbook if provided)
        wb = self._write_plan_performance_xlsx(xlsx_file, headers, rows, workbook)
        print(f"--= Analyzed {len(rows)} unique plans across all test runs")

        return wb

    def _write_plan_performance_xlsx(
        self,
        output_file: str,
        headers: List[str],
        rows: List[List[Any]],
        workbook: Workbook | None = None,
    ) -> Workbook:
        """Write plan performance results to an Excel file with formatting.

        Args:
            output_file: Path to output xlsx file (used if creating new workbook).
            headers: List of header names.
            rows: List of data rows.
            workbook: Optional existing workbook to add sheet to.

        Returns:
            The workbook with plan performance sheet.
        """
        if workbook is None:
            wb = Workbook()
            ws = wb.active
            ws.title = "Plan Performance"
        else:
            wb = workbook
            # Remove default sheet if it exists and is empty
            if "Sheet" in wb.sheetnames:
                del wb["Sheet"]
            ws = wb.create_sheet("Plan Performance")

        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(
            start_color="4472C4", end_color="4472C4", fill_type="solid"
        )
        header_alignment = Alignment(
            horizontal="center", vertical="bottom", wrap_text=False, textRotation=90
        )

        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # Write headers
        for col_idx, header in enumerate(headers, start=1):
            # Add leading space and replace underscores with spaces for readability
            display_header = "        " + header.replace("_", " ")
            cell = ws.cell(row=1, column=col_idx, value=display_header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # Define fonts for coloring
        green_font = Font(color="006400")  # Dark green
        red_font = Font(color="8B0000")  # Dark red

        # Identify column indices
        success_rate_col = (
            headers.index("success_rate") + 1 if "success_rate" in headers else None
        )
        successes_col = (
            headers.index("successes") + 1 if "successes" in headers else None
        )
        failures_col = headers.index("failures") + 1 if "failures" in headers else None

        # Write data rows
        for row_idx, row_data in enumerate(rows, start=2):
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border

                # Format success_rate as percentage
                if col_idx == success_rate_col and isinstance(value, (int, float)):
                    cell.number_format = "0.0%"
                    cell.alignment = Alignment(horizontal="right")
                    # Color based on success rate
                    if value >= 0.8:
                        cell.font = green_font
                    elif value < 0.5:
                        cell.font = red_font
                elif col_idx == successes_col and isinstance(value, (int, float)):
                    cell.font = green_font
                    cell.alignment = Alignment(horizontal="right")
                elif col_idx == failures_col and isinstance(value, (int, float)):
                    if value > 0:
                        cell.font = red_font
                    cell.alignment = Alignment(horizontal="right")
                elif value is None:
                    cell.value = "-"
                    cell.alignment = Alignment(horizontal="center")
                elif isinstance(value, float):
                    cell.number_format = "0.0"
                    cell.alignment = Alignment(horizontal="right")
                elif isinstance(value, int):
                    cell.alignment = Alignment(horizontal="right")

        # Auto-adjust column widths based on original data (before cell transformations)
        # Build set of percentage columns for formatting
        percentage_col_names = {"success_rate"}
        percentage_cols = set()
        for col_idx, header in enumerate(headers, start=1):
            if header in percentage_col_names:
                percentage_cols.add(col_idx)

        for col_idx in range(len(headers)):
            column_letter = get_column_letter(col_idx + 1)
            col_idx_1based = col_idx + 1
            max_length = 0
            for row_data in rows:
                if col_idx < len(row_data):
                    value = row_data[col_idx]
                    if value is not None:
                        # For percentage columns, estimate formatted width (e.g., "85.5%")
                        if col_idx_1based in percentage_cols and isinstance(
                            value, (int, float)
                        ):
                            formatted = f"{value * 100:.1f}%"
                            max_length = max(max_length, len(formatted))
                        # For floats, estimate formatted width (e.g., "0.123")
                        elif isinstance(value, float):
                            formatted = f"{value:.1f}"
                            max_length = max(max_length, len(formatted))
                        else:
                            max_length = max(max_length, len(str(value)))
            # Minimal padding, just enough to fit content
            adjusted_width = max(max_length + 1, 5.5)  # Minimum width of 5.5
            ws.column_dimensions[column_letter].width = adjusted_width

        # Set header row height to accommodate rotated text
        max_header_len = max(len("        " + h.replace("_", " ")) for h in headers)
        header_row_height = max(max_header_len * 7, 100)
        ws.row_dimensions[1].height = header_row_height

        # Freeze the header row
        ws.freeze_panes = "A2"

        # Add autofilter
        ws.auto_filter.ref = ws.dimensions

        return wb

    def generate_variation_summary_report(
        self,
        workbook: Workbook | None = None,
    ) -> Workbook | None:
        """Generate a summary report showing each model's baseline and variation performance.

        Creates a table with one row per model showing:
        - Baseline success %
        - Each variation's success % and delta from baseline

        Args:
            workbook: Optional existing workbook to add sheet to.

        Returns:
            The workbook with variation summary sheet added, or None if no data.
        """
        if workbook is None:
            print("No workbook provided for variation summary.")
            return None

        # Get data from Results sheet
        if "Results" not in workbook.sheetnames:
            print("Results sheet not found in workbook.")
            return workbook

        results_ws = workbook["Results"]

        # Read headers from row 1 (strip whitespace and normalize)
        raw_headers = [cell.value for cell in results_ws[1]]
        headers = [
            h.strip().replace(" ", "_") if isinstance(h, str) else h
            for h in raw_headers
        ]

        # Find column indices (0-indexed for headers list)
        # Headers may have underscores instead of spaces, so normalize lookup
        def find_col(name: str) -> int | None:
            # Try exact match first
            if name in headers:
                return headers.index(name)
            # Try with underscores replaced by spaces
            alt_name = name.replace(" ", "_")
            if alt_name in headers:
                return headers.index(alt_name)
            # Try partial match (header contains the name)
            for i, h in enumerate(headers):
                if isinstance(h, str) and name in h.lower():
                    return i
            return None

        model_col = find_col("model")
        config_col = find_col("config")
        rep_col = find_col("rep")
        test_set_col = find_col("test_set_name")
        success_col = find_col("success_%")
        stdev_col = find_col("success stdev")

        if any(
            c is None
            for c in [model_col, config_col, rep_col, test_set_col, success_col]
        ):
            print(
                f"Required columns not found in Results sheet. Headers: {headers[:10]}"
            )
            print(
                f"  model_col={model_col}, config_col={config_col}, rep_col={rep_col}, test_set_col={test_set_col}, success_col={success_col}"
            )
            return workbook

        # Use module-level BASELINE_CONFIG

        # Parse baseline config to determine which variations to show
        # BASELINE_CONFIG format: T0_Fs_H60_C0_P2_I0_R1_S1_E0_M4096
        def parse_baseline_config(config: str) -> dict[str, str]:
            """Extract parameter values from config string."""
            parts = config.split("_")
            result = {}
            for part in parts:
                if part.startswith("T"):
                    result["T"] = part
                elif part.startswith("F"):
                    result["F"] = part
                elif part.startswith("H"):
                    result["H"] = part
                elif part.startswith("C"):
                    result["C"] = part
                elif part.startswith("P"):
                    result["P"] = part
                elif part.startswith("I"):
                    result["I"] = part
                elif part.startswith("R"):
                    result["R"] = part
                elif part.startswith("S"):
                    result["S"] = part
                elif part.startswith("E"):
                    result["E"] = part
                elif part.startswith("M"):
                    result["M"] = part
            return result

        baseline_params = parse_baseline_config(BASELINE_CONFIG)

        # Define variations dynamically based on baseline
        # Each variation shows options that differ from the baseline
        variations: dict[str, list[tuple[str, str]]] = {}

        # Text only: if baseline is T0, show T1
        if baseline_params.get("T") == "T0":
            variations["text_only"] = [("T1", "text only")]
        else:
            variations["text_only"] = [("T0", "with images")]

        # Feedback type: show options different from baseline
        baseline_f = baseline_params.get("F", "Fs")
        feedback_variations = []
        if baseline_f != "Fn":
            feedback_variations.append(("Fn", "no feedback"))
        if baseline_f != "Fd":
            feedback_variations.append(("Fd", "detailed feedback"))
        if baseline_f != "Fs":
            feedback_variations.append(("Fs", "simple feedback"))
        if feedback_variations:
            variations["feedback_type"] = feedback_variations

        # Hand transparency: if baseline has hand (H60), show no hand (H00)
        baseline_h = baseline_params.get("H", "H60")
        if baseline_h != "H00":
            variations["hand_transparency"] = [("H00", "no hand")]
        else:
            variations["hand_transparency"] = [("H60", "show hand")]

        # Previous image: show options different from baseline
        baseline_i = baseline_params.get("I", "I0")
        img_variations = []
        if baseline_i != "I0":
            img_variations.append(("I0", "no prev img"))
        if baseline_i != "I1":
            img_variations.append(("I1", "prev img"))
        if baseline_i != "I2":
            img_variations.append(("I2", "grayscale img"))
        if img_variations:
            variations["previous_image"] = img_variations

        # Use memory: if baseline is R1 (memory on), show R0; otherwise show R1
        if baseline_params.get("R") == "R1":
            variations["use_memory"] = [("R0", "no memory")]
        else:
            variations["use_memory"] = [("R1", "use memory")]

        # Full steps: if baseline is S1, show S0
        if baseline_params.get("S") == "S1":
            variations["full_steps"] = [("S0", "no full steps")]
        else:
            variations["full_steps"] = [("S1", "full steps")]

        # Build list of all variation columns in order
        variation_columns = []
        for param, var_list in variations.items():
            for code, display_name in var_list:
                variation_columns.append((param, code, display_name))

        # Collect data: model -> {config -> (success %, stdev)}
        # Only look at AVG rows where test_set_name == "ALL"
        model_data: dict[str, dict[str, tuple[float, float | None]]] = {}

        for row_idx, row in enumerate(
            results_ws.iter_rows(min_row=2, values_only=True)
        ):
            if row[rep_col] != "AVG":
                continue
            if row[test_set_col] != "ALL":
                continue

            model_name = row[model_col]
            config_name = row[config_col]
            success_pct = row[success_col]
            stdev_val = row[stdev_col] if stdev_col is not None else None

            if model_name is None or config_name is None:
                continue
            if not isinstance(success_pct, (int, float)) or isinstance(
                success_pct, bool
            ):
                continue

            # Validate stdev
            if not isinstance(stdev_val, (int, float)) or isinstance(stdev_val, bool):
                stdev_val = None

            if model_name not in model_data:
                model_data[model_name] = {}
            model_data[model_name][config_name] = (success_pct, stdev_val)

        if not model_data:
            print("No AVG+ALL data found for variation summary.")
            return workbook

        # Helper function to find a config that varies only the specified code
        def find_variation_config(
            baseline: str, param: str, old_code: str, new_code: str
        ) -> str:
            """Create the config string with a single variation from baseline.

            This ensures we only look at configs that differ from baseline in exactly
            one dimension. Configs that vary in multiple dimensions won't match any
            of the generated variation config strings and will be excluded.
            """
            parts = baseline.split("_")
            new_parts = []
            for part in parts:
                # Find the part that starts with the parameter's prefix
                if param == "text_only" and part.startswith("T"):
                    new_parts.append(new_code)
                elif param == "feedback_type" and part.startswith("F"):
                    new_parts.append(new_code)
                elif param == "hand_transparency" and part.startswith("H"):
                    new_parts.append(new_code)
                elif param == "previous_image" and part.startswith("I"):
                    new_parts.append(new_code)
                elif param == "use_memory" and part.startswith("R"):
                    new_parts.append(new_code)
                elif param == "full_steps" and part.startswith("S"):
                    new_parts.append(new_code)
                else:
                    new_parts.append(part)
            return "_".join(new_parts)

        # Helper function to normalize config strings for comparison (ignore M, E, P, C)
        def normalize_config(cfg: str) -> str:
            """Remove M, E, P, C parts from config string for comparison."""
            if not cfg:
                return ""
            parts = cfg.split("_")
            filtered = [p for p in parts if p and p[0] not in ("M", "E", "P", "C")]
            return "_".join(filtered)

        # Helper function to find config value by normalized key
        def find_config_value(
            configs: dict[str, tuple[float, float | None]], target_config: str
        ) -> tuple[float, float | None] | None:
            """Find a config value, matching by normalized config (ignoring M, E, P, C).
            Returns (success_pct, stdev) tuple or None if not found."""
            target_normalized = normalize_config(target_config)
            for config_key, value in configs.items():
                if normalize_config(config_key) == target_normalized:
                    return value
            return None

        # Build the summary rows with separate columns for success % and delta
        # Each variation gets two columns: success % and % change
        # Also track significance for highlighting (ranges don't overlap)
        summary_rows = []
        # Track significance for each % change cell: row_idx -> col_idx -> is_significant
        significance_for_change_cells: dict[int, dict[int, bool]] = {}

        for model_name in sorted(model_data.keys()):
            configs = model_data[model_name]
            baseline_result = find_config_value(configs, BASELINE_CONFIG)
            baseline_success = baseline_result[0] if baseline_result else None
            baseline_stdev = baseline_result[1] if baseline_result else None

            row_data = [model_name]

            # Baseline column (single column)
            if baseline_success is not None:
                row_data.append(baseline_success)
            else:
                row_data.append("")

            # Variation columns (two columns each: success %, % change)
            # Track significance for this row's % change columns
            row_significance_map: dict[int, bool] = {}
            col_idx = 4  # Start at column 4 (1-indexed: model=1, baseline=2, first variation success=3, first % change=4)

            for param, code, display_name in variation_columns:
                variation_config = find_variation_config(
                    BASELINE_CONFIG, param, "", code
                )
                variation_result = find_config_value(configs, variation_config)
                variation_success = variation_result[0] if variation_result else None
                variation_stdev = variation_result[1] if variation_result else None

                if variation_success is not None:
                    row_data.append(variation_success)  # success %
                    if baseline_success is not None:
                        delta = variation_success - baseline_success
                        row_data.append(delta)  # % change
                        # Check for significance: ranges don't overlap
                        is_significant = False
                        if baseline_stdev is not None and variation_stdev is not None:
                            baseline_lower = baseline_success - baseline_stdev
                            baseline_upper = baseline_success + baseline_stdev
                            variation_lower = variation_success - variation_stdev
                            variation_upper = variation_success + variation_stdev
                            # Significant if ranges don't overlap
                            if (
                                baseline_upper < variation_lower
                                or variation_upper < baseline_lower
                            ):
                                is_significant = True
                        row_significance_map[col_idx] = is_significant
                    else:
                        row_data.append("")  # No baseline to compare
                else:
                    row_data.append("")  # No success %
                    row_data.append("")  # No % change

                col_idx += 2  # Move to next % change column

            summary_rows.append(row_data)
            significance_for_change_cells[len(summary_rows) - 1] = row_significance_map

        # Create the worksheet
        ws = workbook.create_sheet("Variation Summary")

        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(
            start_color="4472C4", end_color="4472C4", fill_type="solid"
        )
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        baseline_fill = PatternFill(
            start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"
        )

        # Write headers with merged cells for variations
        # Row 1: Model | baseline | variation1 (merged 2 cols) | variation2 (merged 2 cols) | ...
        # Row 2: (empty) | (empty) | success % | % change | success % | % change | ...

        # First, write row 1 headers
        ws.cell(row=1, column=1, value="model").font = header_font
        ws.cell(row=1, column=1).fill = header_fill
        ws.cell(row=1, column=1).alignment = header_alignment
        ws.cell(row=1, column=1).border = thin_border

        ws.cell(row=1, column=2, value="baseline").font = header_font
        ws.cell(row=1, column=2).fill = header_fill
        ws.cell(row=1, column=2).alignment = header_alignment
        ws.cell(row=1, column=2).border = thin_border

        # Merge row 1 and 2 for model and baseline columns
        ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
        ws.merge_cells(start_row=1, start_column=2, end_row=2, end_column=2)

        # Write variation headers (merged across 2 columns each)
        col_idx = 3
        for param, code, display_name in variation_columns:
            # Merge two columns for this variation name
            ws.merge_cells(
                start_row=1, start_column=col_idx, end_row=1, end_column=col_idx + 1
            )
            cell = ws.cell(row=1, column=col_idx, value=display_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
            # Also style the merged cell's right side
            ws.cell(row=1, column=col_idx + 1).border = thin_border

            # Row 2: sub-headers for success % and % change
            success_cell = ws.cell(row=2, column=col_idx, value="success %")
            success_cell.font = header_font
            success_cell.fill = header_fill
            success_cell.alignment = header_alignment
            success_cell.border = thin_border

            change_cell = ws.cell(row=2, column=col_idx + 1, value="% change")
            change_cell.font = header_font
            change_cell.fill = header_fill
            change_cell.alignment = header_alignment
            change_cell.border = thin_border

            col_idx += 2

        # Write data rows (starting from row 3 now)
        for row_idx, row_data in enumerate(summary_rows, start=3):
            data_idx = 0

            # Model name (column 1)
            cell = ws.cell(row=row_idx, column=1, value=row_data[data_idx])
            cell.border = thin_border
            data_idx += 1

            # Baseline column (column 2)
            cell = ws.cell(row=row_idx, column=2)
            cell.border = thin_border
            cell.fill = baseline_fill
            value = row_data[data_idx]
            if isinstance(value, (int, float)) and not isinstance(value, bool):
                cell.value = value
                cell.number_format = "0.0%"
                cell.alignment = Alignment(horizontal="center")
            data_idx += 1

            # Variation columns (2 columns each)
            col_idx = 3
            for _ in variation_columns:
                # Success % column
                success_val = row_data[data_idx] if data_idx < len(row_data) else ""
                success_cell = ws.cell(row=row_idx, column=col_idx)
                success_cell.border = thin_border
                success_cell.alignment = Alignment(horizontal="center")
                if isinstance(success_val, (int, float)) and not isinstance(
                    success_val, bool
                ):
                    success_cell.value = success_val
                    success_cell.number_format = "0.0%"
                data_idx += 1

                # % change column
                delta_val = row_data[data_idx] if data_idx < len(row_data) else ""
                delta_cell = ws.cell(row=row_idx, column=col_idx + 1)
                delta_cell.border = thin_border
                delta_cell.alignment = Alignment(horizontal="center")
                if isinstance(delta_val, (int, float)) and not isinstance(
                    delta_val, bool
                ):
                    delta_cell.value = delta_val
                    delta_cell.number_format = "+0.0%;-0.0%;0.0%"
                    # Color code the % change
                    if delta_val > 0.001:
                        delta_cell.font = Font(color="008000")  # Green
                    elif delta_val < -0.001:
                        delta_cell.font = Font(color="FF0000")  # Red
                    # Highlight yellow if ranges don't overlap (significant change)
                    row_significance_map = significance_for_change_cells.get(
                        row_idx - 3, {}
                    )  # row_idx is 1-indexed, data starts at row 3
                    is_significant = row_significance_map.get(
                        col_idx + 1, False
                    )  # col_idx+1 is the % change column
                    if is_significant:
                        delta_cell.fill = PatternFill(
                            start_color="FFFF00",
                            end_color="FFFF00",
                            fill_type="solid",
                        )
                data_idx += 1

                col_idx += 2

        # Calculate total columns: 1 (model) + 1 (baseline) + 2 * num_variations
        total_cols = 2 + 2 * len(variation_columns)

        # Auto-size columns with compact widths
        for col_idx in range(1, total_cols + 1):
            column_letter = get_column_letter(col_idx)
            max_length = 0
            for row in ws.iter_rows(
                min_row=1, max_row=ws.max_row, min_col=col_idx, max_col=col_idx
            ):
                for cell in row:
                    if cell.value:
                        # For percentages, estimate formatted width
                        if isinstance(cell.value, float):
                            formatted = f"{cell.value * 100:.1f}%"
                            max_length = max(max_length, len(formatted))
                        else:
                            max_length = max(max_length, len(str(cell.value)))
            # Compact width: just enough to fit content with minimal padding
            ws.column_dimensions[column_letter].width = max(max_length + 1, 6)

        # Freeze header rows
        ws.freeze_panes = "A3"

        # Add bar chart showing success % for baseline and all variations
        if summary_rows:
            from openpyxl.chart import BarChart, Reference, Series

            chart = BarChart()
            chart.type = "col"  # Column chart (vertical bars)
            chart.style = 10
            chart.title = "Success Rate by Configuration"
            chart.y_axis.title = "Success %"
            chart.y_axis.numFmt = "0%"
            chart.x_axis.title = "Model"

            # Data starts at row 3 (after 2 header rows)
            # Columns: 1=model, 2=baseline, 3=var1 success, 4=var1 change, 5=var2 success, ...
            # We want baseline (col 2) and every odd column after that (3, 5, 7, ...)
            num_data_rows = len(summary_rows)
            data_start_row = 3
            data_end_row = data_start_row + num_data_rows - 1

            # Categories (model names) - column 1
            categories = Reference(
                ws, min_col=1, min_row=data_start_row, max_row=data_end_row
            )

            # Baseline data - column 2 (data only, no header)
            baseline_data = Reference(
                ws, min_col=2, min_row=data_start_row, max_row=data_end_row
            )
            baseline_series = Series(baseline_data, title="baseline")
            chart.series.append(baseline_series)

            # Variation success % columns (every other column starting from 3)
            col_idx = 3
            for param, code, display_name in variation_columns:
                var_data = Reference(
                    ws, min_col=col_idx, min_row=data_start_row, max_row=data_end_row
                )
                var_series = Series(var_data, title=display_name)
                chart.series.append(var_series)
                col_idx += 2  # Skip the % change column

            chart.set_categories(categories)
            chart.shape = 4  # Rectangle bars
            chart.width = 20  # Chart width in cm
            chart.height = 12  # Chart height in cm

            # Position chart below the data
            chart_row = data_end_row + 3
            ws.add_chart(chart, f"A{chart_row}")

        print(
            f"Generated variation summary with {len(summary_rows)} models and {len(variation_columns)} variations."
        )
        return workbook

    def generate_config_analysis_report(
        self,
        workbook: Workbook | None = None,
    ) -> Workbook | None:
        """Generate a report analyzing success rates by configuration dimension.

        For each model and configuration option, shows success rates when that
        config is True vs False (or for feedback_type: none/simple/detailed).

        Args:
            workbook: Optional existing workbook to add sheet to.

        Returns:
            The workbook with config analysis sheet added, or None if no data.
        """
        from collections import defaultdict

        # Config dimensions to analyze
        config_dimensions = [
            "text_only",
            "feedback_type",
            "hand_transparency",
            "include_common_sense",
            "previous_image",
            "use_memory",
        ]

        # Collect all results grouped by model name and config values
        # Structure: {model_name: {config_key: {config_value: [success_count, fail_count]}}}
        model_config_stats: dict[str, dict[str, dict[Any, list[int]]]] = defaultdict(
            lambda: defaultdict(lambda: defaultdict(lambda: [0, 0]))
        )

        if not os.path.exists(self.test_dir):
            print(f"Test directory not found: {self.test_dir}")
            return workbook

        # Scan all test directories
        test_names = [
            t
            for t in os.listdir(self.test_dir)
            if os.path.isdir(os.path.join(self.test_dir, t))
        ]

        print(f"Scanning {len(test_names)} test directories for config analysis...")

        for test_name in test_names:
            test_path = os.path.join(self.test_dir, test_name)
            if not os.path.isdir(test_path):
                continue

            for model_dir in os.listdir(test_path):
                model_path = os.path.join(test_path, model_dir)
                if not os.path.isdir(model_path):
                    continue

                config_file = os.path.join(model_path, "config.json")
                test_results_file = os.path.join(model_path, "test_results.json")

                if not os.path.exists(config_file) or not os.path.exists(
                    test_results_file
                ):
                    continue

                # Load and normalize config
                config_data = self._read_json_cached(config_file)
                config_data = EvaluationConfig.normalize_config_dict(config_data)

                # Get model name from model_path in config (e.g., "gpt-4o")
                model_name = config_data.get("model_path", "")
                if not model_name:
                    # Fallback to model_name field
                    model_name = config_data.get("model_name", model_dir)

                # Load test results
                test_results_data = self._read_json_cached(test_results_file)

                # Count successes and failures
                results_list = test_results_data.get("test_results", [])
                for result in results_list:
                    task_failed = result.get("task_failed", False)

                    # For each config dimension, record the result
                    for config_key in config_dimensions:
                        config_value = config_data.get(config_key)
                        if config_value is not None:
                            # Normalize hand_transparency to boolean-like
                            if config_key == "hand_transparency":
                                config_value = config_value > 0  # True if non-zero

                            if task_failed:
                                model_config_stats[model_name][config_key][
                                    config_value
                                ][1] += 1
                            else:
                                model_config_stats[model_name][config_key][
                                    config_value
                                ][0] += 1

        if not model_config_stats:
            print("No config data found to analyze.")
            return workbook

        # Debug: print aggregated stats
        print(f"Config analysis found {len(model_config_stats)} unique models:")
        for model_name in sorted(model_config_stats.keys()):
            print(f"  - {model_name}")
            for config_key, value_stats in model_config_stats[model_name].items():
                stats_summary = ", ".join(
                    f"{k}={v[0]}s/{v[1]}f" for k, v in value_stats.items()
                )
                print(f"      {config_key}: {stats_summary}")

        # Build output data
        # Headers: Model | Config | Success% x3 | Values x3 | Fail% x3 | Values x3
        headers = [
            "model",
            "config",
            "success_pct_1",
            "success_pct_2",
            "success_pct_3",
            "success_val_1",
            "success_val_2",
            "success_val_3",
            "fail_pct_1",
            "fail_pct_2",
            "fail_pct_3",
            "fail_val_1",
            "fail_val_2",
            "fail_val_3",
        ]

        # Display names for config dimensions
        config_display_names = {
            "text_only": "text only",
            "feedback_type": "feedback type",
            "hand_transparency": "show hand",
            "include_common_sense": "common sense",
            "previous_image": "prev image",
            "use_memory": "use memory",
        }

        rows: List[List[Any]] = []
        for model_name in sorted(model_config_stats.keys()):
            config_stats = model_config_stats[model_name]

            for config_key in config_dimensions:
                if config_key not in config_stats:
                    continue

                display_name = config_display_names.get(config_key, config_key)
                value_stats = config_stats[config_key]

                if config_key == "feedback_type":
                    # Special handling for feedback_type (3 values: none, simple, detailed)
                    none_stats = value_stats.get("none", [0, 0])
                    simple_stats = value_stats.get("simple", [0, 0])
                    detailed_stats = value_stats.get("detailed", [0, 0])

                    none_total = none_stats[0] + none_stats[1]
                    simple_total = simple_stats[0] + simple_stats[1]
                    detailed_total = detailed_stats[0] + detailed_stats[1]

                    # Success percentages
                    none_success = (
                        none_stats[0] / none_total if none_total > 0 else None
                    )
                    simple_success = (
                        simple_stats[0] / simple_total if simple_total > 0 else None
                    )
                    detailed_success = (
                        detailed_stats[0] / detailed_total
                        if detailed_total > 0
                        else None
                    )

                    # Fail percentages
                    none_fail = none_stats[1] / none_total if none_total > 0 else None
                    simple_fail = (
                        simple_stats[1] / simple_total if simple_total > 0 else None
                    )
                    detailed_fail = (
                        detailed_stats[1] / detailed_total
                        if detailed_total > 0
                        else None
                    )

                    rows.append(
                        [
                            model_name,
                            display_name,
                            none_success,
                            simple_success,
                            detailed_success,
                            "None",
                            "Simple",
                            "Detailed",
                            none_fail,
                            simple_fail,
                            detailed_fail,
                            "None",
                            "Simple",
                            "Detailed",
                        ]
                    )
                else:
                    # Boolean-like config (False, True)
                    false_stats = value_stats.get(False, [0, 0])
                    true_stats = value_stats.get(True, [0, 0])

                    false_total = false_stats[0] + false_stats[1]
                    true_total = true_stats[0] + true_stats[1]

                    # Success percentages
                    false_success = (
                        false_stats[0] / false_total if false_total > 0 else None
                    )
                    true_success = (
                        true_stats[0] / true_total if true_total > 0 else None
                    )

                    # Fail percentages
                    false_fail = (
                        false_stats[1] / false_total if false_total > 0 else None
                    )
                    true_fail = true_stats[1] / true_total if true_total > 0 else None

                    rows.append(
                        [
                            model_name,
                            display_name,
                            false_success,
                            true_success,
                            None,
                            "False",
                            "True",
                            None,
                            false_fail,
                            true_fail,
                            None,
                            "False",
                            "True",
                            None,
                        ]
                    )

        # Write to Excel
        wb = self._write_config_analysis_xlsx(headers, rows, workbook)
        print(f"--= Generated config analysis with {len(rows)} rows")

        return wb

    def _write_config_analysis_xlsx(
        self,
        headers: List[str],
        rows: List[List[Any]],
        workbook: Workbook | None = None,
    ) -> Workbook:
        """Write config analysis results to an Excel worksheet.

        Args:
            headers: List of header names.
            rows: List of data rows.
            workbook: Optional existing workbook to add sheet to.

        Returns:
            The workbook with config analysis sheet.
        """
        if workbook is None:
            wb = Workbook()
            ws = wb.active
            ws.title = "Config Analysis"
        else:
            wb = workbook
            if "Sheet" in wb.sheetnames:
                del wb["Sheet"]
            ws = wb.create_sheet("Config Analysis")

        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(
            start_color="4472C4", end_color="4472C4", fill_type="solid"
        )
        header_alignment = Alignment(
            horizontal="center", vertical="bottom", wrap_text=False, textRotation=90
        )

        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # Custom header display names
        # Columns: Model | Config | Success% x3 | Values x3 | Fail% x3 | Values x3
        header_display = {
            "model": "Model",
            "config": "Config",
            "success_pct_1": "% Success",
            "success_pct_2": "% Success",
            "success_pct_3": "% Success",
            "success_val_1": "Value",
            "success_val_2": "Value",
            "success_val_3": "Value",
            "fail_pct_1": "% Fail",
            "fail_pct_2": "% Fail",
            "fail_pct_3": "% Fail",
            "fail_val_1": "Value",
            "fail_val_2": "Value",
            "fail_val_3": "Value",
        }

        # Write headers
        for col_idx, header in enumerate(headers, start=1):
            display_header = header_display.get(header, header)
            cell = ws.cell(row=1, column=col_idx, value=display_header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # Define fonts for coloring
        green_font = Font(bold=True, color="006400")  # Dark green
        red_font = Font(bold=True, color="8B0000")  # Dark red

        # Column indices (1-based):
        # 1=model, 2=config
        # 3-5=success_pct, 6-8=success_val
        # 9-11=fail_pct, 12-14=fail_val
        success_pct_columns = {3, 4, 5}
        success_val_columns = {6, 7, 8}
        fail_pct_columns = {9, 10, 11}
        fail_val_columns = {12, 13, 14}

        # Write data rows
        for row_idx, row_data in enumerate(rows, start=2):
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border

                if col_idx in success_pct_columns:
                    # Format success percentage columns (green)
                    if isinstance(value, (int, float)):
                        cell.number_format = "0.0%"
                        cell.alignment = Alignment(horizontal="right")
                        cell.font = green_font
                    elif value is None:
                        cell.value = "-"
                        cell.alignment = Alignment(horizontal="center")
                elif col_idx in fail_pct_columns:
                    # Format fail percentage columns (red)
                    if isinstance(value, (int, float)):
                        cell.number_format = "0.0%"
                        cell.alignment = Alignment(horizontal="right")
                        cell.font = red_font
                    elif value is None:
                        cell.value = "-"
                        cell.alignment = Alignment(horizontal="center")
                elif col_idx in success_val_columns or col_idx in fail_val_columns:
                    # Value label columns
                    if value is None:
                        cell.value = "-"
                    cell.alignment = Alignment(horizontal="center")
                else:
                    cell.alignment = Alignment(horizontal="left")

        # Auto-adjust column widths
        for col_idx, header in enumerate(headers, start=1):
            column_letter = get_column_letter(col_idx)
            if header == "model":
                # Model names can be long
                max_length = max(
                    (
                        len(str(row[col_idx - 1]))
                        for row in rows
                        if col_idx - 1 < len(row)
                    ),
                    default=10,
                )
                ws.column_dimensions[column_letter].width = max(max_length + 2, 15)
            elif header == "config":
                ws.column_dimensions[column_letter].width = 15
            elif "pct" in header:
                ws.column_dimensions[column_letter].width = 10
            elif "val" in header:
                ws.column_dimensions[column_letter].width = 9

        # Set header row height to accommodate rotated text
        max_header_len = max(len(header_display.get(h, h)) for h in headers)
        header_row_height = max(max_header_len * 7, 80)
        ws.row_dimensions[1].height = header_row_height

        # Freeze the header row
        ws.freeze_panes = "A2"

        # Add autofilter
        ws.auto_filter.ref = ws.dimensions

        return wb

    def generate_error_analysis_report(
        self,
        workbook: Workbook | None = None,
    ) -> Workbook | None:
        """Generate a report analyzing errors across all plans, grouped by model.

        Shows all errors in order of frequency for each model, with the action,
        normalized error message, and placeholder object distributions.

        Args:
            workbook: Optional existing workbook to add sheet to.

        Returns:
            The workbook with error analysis sheet added, or None if no data.
        """
        import re
        from collections import defaultdict
        from dataclasses import dataclass, field

        # Structure: {model_name: {full_error: {count, placeholder_values}}}
        # Where placeholder_values is: {placeholder -> {object_type -> count}}
        @dataclass
        class ErrorInfo:
            count: int = 0
            placeholder_values: dict = field(default_factory=dict)

        model_errors: dict[str, dict[str, ErrorInfo]] = defaultdict(
            lambda: defaultdict(ErrorInfo)
        )
        model_total_errors: dict[str, int] = defaultdict(int)

        if not os.path.exists(self.test_dir):
            print(f"Test directory not found: {self.test_dir}")
            return workbook

        # Words to exclude from object matching (common English words)
        excluded_words = {
            "Cannot",
            "The",
            "This",
            "That",
            "There",
            "Here",
            "What",
            "When",
            "Where",
            "Which",
            "While",
            "With",
            "Without",
            "Could",
            "Would",
            "Should",
            "Must",
            "Will",
            "Shall",
            "Have",
            "Has",
            "Had",
            "Does",
            "Did",
            "Done",
            "Been",
            "Being",
            "Was",
            "Were",
            "Are",
            "Is",
            "Am",
            "Be",
            "Not",
            "But",
            "And",
            "For",
            "From",
            "Into",
            "Onto",
            "Over",
            "Under",
            "Above",
            "Below",
            "Between",
            "Among",
            "Through",
            "During",
            "Before",
            "After",
            "Since",
            "Until",
            "Already",
            "Also",
            "Always",
            "Never",
            "Ever",
            "Only",
            "Just",
            "Still",
            "Even",
            "Very",
            "Too",
            "Much",
            "Many",
            "Some",
            "Any",
            "All",
            "Each",
            "Every",
            "Both",
            "Either",
            "Neither",
            "Other",
            "Another",
            "Such",
            "Same",
            "Different",
            "PUT",
            "GET",
            "Ran",
            "Failed",
            "Error",
            "Invalid",
            "Missing",
            "Expected",
            "Unexpected",
            "Unable",
            "Slice",
            "Sliced",
            "Types",
            "Type",
            "Action",
            "Object",
            "Objects",
            "Specifier",
            "Candidate",
            "Candidates",
            "None",
            "True",
            "False",
        }

        # Pattern to match object names with suffixes
        object_with_suffix_pattern = re.compile(
            r"[A-Z][a-zA-Z]*(?:_[a-zA-Z0-9]+)+(?:\(Clone\))?"
        )
        # Pattern to match simple CamelCase object names (e.g., WineBottle, SinkBasin)
        simple_object_pattern = re.compile(r"\b([A-Z][a-z]+(?:[A-Z][a-z]+)+)\b")

        def extract_object_type(obj_name: str) -> str:
            """Extract the base object type from a full object name."""
            name = obj_name.replace("(Clone)", "")
            parts = name.split("_")
            type_parts = []
            for part in parts:
                if re.match(r"^[a-f0-9]{6,}$", part, re.IGNORECASE):
                    break
                if re.match(r"^\d+$", part):
                    break
                type_parts.append(part)
            return "_".join(type_parts) if type_parts else parts[0]

        def find_objects_in_message(msg: str) -> list:
            """Find all object references in an error message."""
            result = []
            # First find objects with suffixes (e.g., Bread_fe4bb3e3)
            for match in object_with_suffix_pattern.findall(msg):
                if match not in result:
                    result.append(match)
            # Find simple CamelCase objects (e.g., WineBottle, SinkBasin)
            for match in simple_object_pattern.findall(msg):
                if match not in excluded_words and match not in result:
                    # Check it's not already covered by a suffixed version
                    if not any(m.startswith(match + "_") for m in result):
                        result.append(match)
            # Also find single capitalized words that might be objects
            # (e.g., Pencil, Apple, Bread)
            single_word_pattern = re.compile(r"\b([A-Z][a-z]{2,})\b")
            for match in single_word_pattern.findall(msg):
                if match not in excluded_words and match not in result:
                    if not any(m.startswith(match + "_") for m in result):
                        result.append(match)
            return result

        # Scan all test directories
        test_names = [
            t
            for t in os.listdir(self.test_dir)
            if os.path.isdir(os.path.join(self.test_dir, t))
        ]

        print(f"Scanning {len(test_names)} test directories for error analysis...")

        for test_name in test_names:
            test_path = os.path.join(self.test_dir, test_name)
            if not os.path.isdir(test_path):
                continue

            for model_dir in os.listdir(test_path):
                model_path = os.path.join(test_path, model_dir)
                if not os.path.isdir(model_path):
                    continue

                config_file = os.path.join(model_path, "config.json")
                test_results_file = os.path.join(model_path, "test_results.json")

                if not os.path.exists(config_file) or not os.path.exists(
                    test_results_file
                ):
                    continue

                # Load config to get model name
                config_data = self._read_json_cached(config_file)

                # Get model name from model_path in config
                model_name = config_data.get("model_path", "")
                if not model_name:
                    model_name = config_data.get("model_name", model_dir)

                # Load test results
                test_results_data = self._read_json_cached(test_results_file)

                # Process errors from each result
                results_list = test_results_data.get("test_results", []) or []
                for result in results_list:
                    step_errors = result.get("step_errors") or []
                    for error in step_errors:
                        action_name = error.get("action_name", "")
                        error_msg = error.get("error_msg", "")
                        if not error_msg:
                            continue

                        # Find all unique object names in the error message
                        matches = find_objects_in_message(error_msg)
                        unique_objects = []
                        for m in matches:
                            if m not in unique_objects:
                                unique_objects.append(m)

                        # Replace each unique object with {X}, {Y}, {Z}, etc.
                        placeholder_names = ["X", "Y", "Z", "W", "V"]
                        normalized_msg = error_msg
                        object_types = {}  # placeholder_name -> object_type
                        for i, obj_name in enumerate(unique_objects):
                            ph_name = (
                                placeholder_names[i]
                                if i < len(placeholder_names)
                                else f"O{i}"
                            )
                            placeholder = "{" + ph_name + "}"
                            normalized_msg = normalized_msg.replace(
                                obj_name, placeholder
                            )
                            object_types[ph_name] = extract_object_type(obj_name)

                        # Format: "action_name : normalized_error_msg"
                        full_error = (
                            f"{action_name} : {normalized_msg}"
                            if action_name
                            else normalized_msg
                        )

                        # Update error info
                        error_info = model_errors[model_name][full_error]
                        error_info.count += 1
                        model_total_errors[model_name] += 1

                        # Track placeholder values
                        for ph_name, obj_type in object_types.items():
                            if ph_name not in error_info.placeholder_values:
                                error_info.placeholder_values[ph_name] = {}
                            error_info.placeholder_values[ph_name][obj_type] = (
                                error_info.placeholder_values[ph_name].get(obj_type, 0)
                                + 1
                            )

        if not model_errors:
            print("No error data found to analyze.")
            return workbook

        # Helper to format placeholder values
        def format_placeholder_values(ph_dict: dict) -> str:
            if not ph_dict:
                return ""
            total = sum(ph_dict.values())
            sorted_items = sorted(ph_dict.items(), key=lambda x: x[1], reverse=True)
            parts = [
                f"{obj_type} ({count/total*100:.0f}%)"
                for obj_type, count in sorted_items
            ]
            return ", ".join(parts)

        # Build output data
        # Headers: Model | % | Action | Error Message | {X} | {Y} | {Z}
        headers = ["model", "pct", "action", "error_message", "X", "Y", "Z"]

        rows: List[List[Any]] = []
        for model_name in sorted(model_errors.keys()):
            errors = model_errors[model_name]
            total = model_total_errors[model_name]

            # Sort errors by count descending
            sorted_errors = sorted(
                errors.items(), key=lambda x: x[1].count, reverse=True
            )

            for full_error, error_info in sorted_errors:
                pct = error_info.count / total if total > 0 else 0

                # Parse "action : message" format
                if " : " in full_error:
                    action, message = full_error.split(" : ", 1)
                else:
                    action = ""
                    message = full_error

                # Get placeholder values
                x_values = format_placeholder_values(
                    error_info.placeholder_values.get("X", {})
                )
                y_values = format_placeholder_values(
                    error_info.placeholder_values.get("Y", {})
                )
                z_values = format_placeholder_values(
                    error_info.placeholder_values.get("Z", {})
                )

                rows.append(
                    [model_name, pct, action, message, x_values, y_values, z_values]
                )

        # Write to Excel
        wb = self._write_error_analysis_xlsx(headers, rows, workbook)
        print(
            f"--= Generated error analysis with {len(rows)} rows across {len(model_errors)} models"
        )

        return wb

    def _write_error_analysis_xlsx(
        self,
        headers: List[str],
        rows: List[List[Any]],
        workbook: Workbook | None = None,
    ) -> Workbook:
        """Write error analysis results to an Excel worksheet.

        Args:
            headers: List of header names.
            rows: List of data rows.
            workbook: Optional existing workbook to add sheet to.

        Returns:
            The workbook with error analysis sheet.
        """
        if workbook is None:
            wb = Workbook()
            ws = wb.active
            ws.title = "Error Analysis"
        else:
            wb = workbook
            if "Sheet" in wb.sheetnames:
                del wb["Sheet"]
            ws = wb.create_sheet("Error Analysis")

        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(
            start_color="4472C4", end_color="4472C4", fill_type="solid"
        )
        header_alignment = Alignment(
            horizontal="center", vertical="bottom", wrap_text=False, textRotation=90
        )

        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # Custom header display names
        header_display = {
            "model": "Model",
            "pct": "%",
            "action": "Action",
            "error_message": "Error Message",
            "X": "{X}",
            "Y": "{Y}",
            "Z": "{Z}",
        }

        # Write headers
        for col_idx, header in enumerate(headers, start=1):
            display_header = header_display.get(header, header)
            cell = ws.cell(row=1, column=col_idx, value=display_header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # Define fonts for coloring
        red_font = Font(bold=True, color="8B0000")  # Dark red for percentage

        # Track current model for alternating backgrounds
        light_fill = PatternFill(
            start_color="E6F3FF", end_color="E6F3FF", fill_type="solid"
        )
        current_model = None
        use_light_fill = False

        # Write data rows
        for row_idx, row_data in enumerate(rows, start=2):
            model_name = row_data[0]

            # Alternate background per model
            if model_name != current_model:
                current_model = model_name
                use_light_fill = not use_light_fill

            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border

                # Apply alternating fill
                if use_light_fill:
                    cell.fill = light_fill

                # Format percentage column (column 2)
                if col_idx == 2 and isinstance(value, (int, float)):
                    cell.number_format = "0.0%"
                    cell.alignment = Alignment(horizontal="right")
                    cell.font = red_font
                elif col_idx == 1:
                    cell.alignment = Alignment(horizontal="left")
                elif col_idx == 3:  # Action
                    cell.alignment = Alignment(horizontal="center")
                else:
                    cell.alignment = Alignment(horizontal="left")

        # Auto-adjust column widths
        for col_idx, header in enumerate(headers, start=1):
            column_letter = get_column_letter(col_idx)
            if header == "model":
                max_length = max(
                    (
                        len(str(row[col_idx - 1]))
                        for row in rows
                        if col_idx - 1 < len(row)
                    ),
                    default=10,
                )
                ws.column_dimensions[column_letter].width = max(max_length + 2, 15)
            elif header == "pct":
                ws.column_dimensions[column_letter].width = 8
            elif header == "action":
                ws.column_dimensions[column_letter].width = 12
            elif header == "error_message":
                ws.column_dimensions[column_letter].width = 60
            elif header in ("X", "Y", "Z"):
                ws.column_dimensions[column_letter].width = 35

        # Freeze the header row
        ws.freeze_panes = "A2"

        # Set header row height to accommodate rotated text
        max_header_len = max(len(header_display.get(h, h)) for h in headers)
        header_row_height = max(max_header_len * 7, 80)
        ws.row_dimensions[1].height = header_row_height

        # Add autofilter
        ws.auto_filter.ref = ws.dimensions

        return wb


def main():
    parser = argparse.ArgumentParser(
        description="Print aggregated test results from the test directory"
    )
    parser.add_argument(
        "--test-dir",
        type=str,
        default=None,
        help=f"Directory containing test results (default: {c.TEST_DIR})",
    )
    parser.add_argument(
        "--output",
        "-o",
        type=str,
        default=None,
        help="Output CSV file path (default: {test_dir}/results.csv)",
    )
    parser.add_argument(
        "--models",
        type=str,
        nargs="*",
        default=None,
        help="Specific model directory names to include (default: auto-discover all)",
    )
    parser.add_argument(
        "--model-filter",
        type=str,
        default=None,
        help="Filter models by substring match (e.g., 'google__gemini-3-pro-preview-high')",
    )
    parser.add_argument(
        "--tests",
        type=str,
        nargs="*",
        default=None,
        help="Specific test names to include (default: auto-discover all)",
    )
    parser.add_argument(
        "--skip-plan-performance",
        action="store_true",
        help="Skip generating the plan performance report",
    )

    args = parser.parse_args()

    try:
        printer = ResultsPrinter(test_dir=args.test_dir)

        # If model_filter is specified, discover models and filter them
        model_dirs = args.models
        if args.model_filter:
            discovered_model_dirs, _ = printer.discover_tests_and_models()
            filtered_models = [
                m for m in discovered_model_dirs if args.model_filter in m
            ]
            print(
                f"Model filter '{args.model_filter}' matched {len(filtered_models)} model(s)"
            )
            if model_dirs:
                # Combine with explicit --models (intersection)
                model_dirs = [m for m in model_dirs if m in filtered_models]
            else:
                model_dirs = filtered_models

        result = printer.print_all_results(
            model_dirs=model_dirs,
            test_list=args.tests,
            output_file=args.output,
        )

        if result is None:
            return

        wb, xlsx_file = result

        # Generate plan performance report (adds to same workbook)
        if not args.skip_plan_performance:
            wb = printer.generate_plan_performance_report(workbook=wb) or wb

        # Generate variation summary report (adds to same workbook)
        wb = printer.generate_variation_summary_report(workbook=wb) or wb

        # Generate config analysis report (adds to same workbook)
        wb = printer.generate_config_analysis_report(workbook=wb) or wb

        # Generate error analysis report (adds to same workbook)
        wb = printer.generate_error_analysis_report(workbook=wb) or wb

        # Save the combined workbook
        wb.save(xlsx_file)
        print(f"--= Wrote Excel results to {xlsx_file}")

    except FileReadError as e:
        print_error(str(e))
        sys.exit(1)


if __name__ == "__main__":
    main()
